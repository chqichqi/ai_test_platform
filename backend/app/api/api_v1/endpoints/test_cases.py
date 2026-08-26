"""
测试用例管理API端点
对应需求文档 3.3.3 功能测试用例生成
"""

import json
import io
import csv
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.core.database import get_db
from app.core.models.requirement import TestCase, TestCaseStatus, TestCasePriority, TestCaseType, ExecutionType
from app.core.models.project import Project, Version
from app.core.schemas.requirement import (
    TestCaseCreate, TestCaseUpdate, TestCaseResponse, TestCaseListResponse
)
from app.core.logger import logger
from app.core.middleware.permission_middleware import Permissions

router = APIRouter()


@router.post("/", response_model=TestCaseResponse, status_code=status.HTTP_201_CREATED)
def create_test_case(
    case_in: TestCaseCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """
    创建测试用例
    
    功能要求（需求文档3.3.3）:
    - 手动创建测试用例
    - 支持测试步骤、测试数据
    """
    project = db.query(Project).filter(Project.id == case_in.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail=f"项目ID {case_in.project_id} 不存在")
    
    test_case = TestCase(
        project_id=case_in.project_id,
        version_id=case_in.version_id,
        module=case_in.module,
        name=case_in.name,
        description=case_in.description,
        preconditions=case_in.preconditions,
        test_steps=[s.model_dump() for s in case_in.test_steps] if case_in.test_steps else None,
        expected_result=case_in.expected_result,
        test_data=case_in.test_data,
        priority=case_in.priority,
        case_type=case_in.case_type,
        execution_type=case_in.execution_type,
        tags=case_in.tags,
        generated_by="manual",
        created_by=current_user["user"].id
    )
    
    db.add(test_case)
    db.commit()
    db.refresh(test_case)

    # 方案B：新建用例逻辑=物理（logical_case_id=自身id）
    if not test_case.logical_case_id:
        test_case.logical_case_id = test_case.id
        db.commit()

    logger.info(f"创建测试用例: {test_case.name}")

    return TestCaseResponse.model_validate(test_case)


@router.get("/", response_model=TestCaseListResponse)
def list_test_cases(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    project_id: Optional[int] = Query(None),
    version_id: Optional[int] = Query(None),
    module: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    source: Optional[str] = Query(None, description="来源: ai_generated / imported_req / imported_cases"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """测试用例列表（方案B：带 version_id 时展示该版本视角的【生效行】——
    跨版本派生链 version_id<=V 的非冻结最大修订，而不是物理 version_id 精确过滤）"""
    query = db.query(TestCase)

    if version_id:
        from app.core.models.project import Version as VersionModel
        from app.core.services.case_versioning import resolve_effective_cases
        _v = db.query(VersionModel).filter(VersionModel.id == version_id).first()
        if _v:
            eff_ids = [r.id for r in resolve_effective_cases(db, _v.project_id, version_id)]
            query = query.filter(TestCase.id.in_(eff_ids))
        else:
            query = query.filter(TestCase.version_id == version_id)  # 版本不存在兜底
    elif project_id:
        query = query.filter(TestCase.project_id == project_id)

    if module:
        query = query.filter(TestCase.module == module)
    
    if priority:
        query = query.filter(TestCase.priority == priority)
    
    if status:
        query = query.filter(TestCase.status == status)

    if source:
        source_map = {
            "ai": "ai",
            "ai_generated": "ai",
            "imported_req": "ai",
            "business_flow": "business_flow",
            "imported_cases": "manual",
        }
        query = query.filter(TestCase.generated_by == source_map.get(source, source))

    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                TestCase.name.ilike(pattern),
                TestCase.description.ilike(pattern)
            )
        )
    
    total = query.count()
    total_pages = (total + page_size - 1) // page_size
    
    cases = query.order_by(TestCase.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return TestCaseListResponse(
        items=[TestCaseResponse.model_validate(c) for c in cases],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.get("/{case_id}", response_model=TestCaseResponse)
def get_test_case(
    case_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取测试用例详情"""
    test_case = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"测试用例ID {case_id} 不存在")
    return TestCaseResponse.model_validate(test_case)


@router.put("/{case_id}", response_model=TestCaseResponse)
def update_test_case(
    case_id: int,
    case_in: TestCaseUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_UPDATE)
):
    """更新测试用例"""
    test_case = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"测试用例ID {case_id} 不存在")
    
    update_data = case_in.model_dump(exclude_unset=True)
    
    if 'test_steps' in update_data and update_data['test_steps']:
        update_data['test_steps'] = [s.model_dump() if hasattr(s, 'model_dump') else s for s in update_data['test_steps']]
    
    for field, value in update_data.items():
        setattr(test_case, field, value)
    
    db.commit()
    db.refresh(test_case)
    
    logger.info(f"更新测试用例: {test_case.name}")
    
    return TestCaseResponse.model_validate(test_case)


@router.delete("/{case_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_test_case(
    case_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_DELETE)
):
    """删除测试用例（有UI用例依赖时阻止删除）"""
    test_case = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"测试用例ID {case_id} 不存在")

    # 登录模块功能用例保护（平台内部约定名）：业务流导入时自动生成（登录链路
    # 三件套的组成部分），与 __login__ UI 用例同为系统资产——删掉即登录链路断链，
    # 与 UI 用例一致：导入后不可删除（2026-08-23 用户反馈误删）
    if (getattr(test_case, 'module', '') or '').strip() == '登录模块':
        raise HTTPException(
            status_code=409,
            detail=f"登录模块功能用例由系统导入自动生成，不可删除"
        )

    # 检查是否存在关联的 WebUI 用例（软删的不算依赖）——方案B 按逻辑 id 匹配（含历史物理绑定）
    from app.core.models.web_ui_test import WebUITestCase as WebUITestCaseModel
    _logical = test_case.logical_case_id or test_case.id
    _phys_ids = [r.id for r in db.query(TestCase).filter(
        TestCase.logical_case_id == _logical
    ).all()]
    _candidates = list(dict.fromkeys([str(_logical)] + [str(p) for p in _phys_ids]))
    ui_case = db.query(WebUITestCaseModel).filter(
        WebUITestCaseModel.test_case_id.in_(_candidates),
        WebUITestCaseModel.deleted_at.is_(None)
    ).first()
    if ui_case:
        raise HTTPException(
            status_code=409,
            detail=f"该用例已转化为UI用例，请先从「UI用例」页面删除对应的UI用例后再删除功能用例"
        )

    db.delete(test_case)
    db.commit()

    logger.info(f"删除测试用例: {test_case.name}")

    return None


@router.post("/batch-delete")
def batch_delete_test_cases(
    request: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_DELETE)
):
    """批量删除测试用例（有UI用例依赖的跳过并报告）"""
    case_ids = request.get("case_ids", [])
    if not case_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的用例")

    from app.core.models.web_ui_test import WebUITestCase as WebUITestCaseModel

    deleted = 0
    blocked = []
    for cid in case_ids:
        tc = db.query(TestCase).filter(TestCase.id == int(cid)).first()
        if not tc:
            continue

        # 登录模块功能用例保护（同单条删除）：系统导入自动生成，不可删除
        if (getattr(tc, 'module', '') or '').strip() == '登录模块':
            blocked.append({"id": int(cid), "name": tc.name, "reason": "登录模块用例不可删除"})
            continue

        # 检查 UI 用例依赖（软删的不算依赖）——方案B 按逻辑 id 匹配（含历史物理绑定）
        _logical = tc.logical_case_id or tc.id
        _phys_ids = [r.id for r in db.query(TestCase).filter(
            TestCase.logical_case_id == _logical
        ).all()]
        _candidates = list(dict.fromkeys([str(_logical)] + [str(p) for p in _phys_ids]))
        ui_case = db.query(WebUITestCaseModel).filter(
            WebUITestCaseModel.test_case_id.in_(_candidates),
            WebUITestCaseModel.deleted_at.is_(None)
        ).first()
        if ui_case:
            blocked.append({"id": int(cid), "name": tc.name, "reason": "已转化为UI用例"})
            continue

        db.delete(tc)
        deleted += 1

    db.commit()

    logger.info(f"批量删除测试用例: 成功 {deleted}, 阻止 {len(blocked)}")

    _msg = f"成功删除 {deleted} 条"
    if blocked:
        _reasons = {}
        for _b in blocked:
            _r = _b.get("reason") or "无法删除"
            _reasons.setdefault(_r, []).append(_b["name"])
        _msg += "，" + "；".join(
            f"{len(v)} 条{_r}: {', '.join(v[:5])}" for _r, v in _reasons.items()
        )

    return {
        "success": True,
        "deleted": deleted,
        "blocked": blocked,
        "message": _msg,
    }


@router.post("/{case_id}/review")
def review_test_case(
    case_id: int,
    req: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_UPDATE)
):
    """
    审核测试用例

    功能要求（需求文档3.3.3）:
    - 生成的用例需要人工审核
    - 支持 action: submit_for_review / approve / reject
    """
    test_case = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"测试用例ID {case_id} 不存在")

    action = req.get("action", "approve")
    comment = req.get("comment")

    if action == "submit_for_review":
        test_case.status = TestCaseStatus.PENDING_REVIEW.value
    elif action == "approve":
        test_case.status = TestCaseStatus.APPROVED.value
    elif action == "reject":
        test_case.status = TestCaseStatus.REJECTED.value
    else:
        raise HTTPException(status_code=400, detail=f"不支持的操作: {action}")

    test_case.reviewer_id = current_user["user"].id
    test_case.review_comment = comment

    from datetime import datetime
    test_case.reviewed_at = datetime.utcnow()

    db.commit()
    db.refresh(test_case)

    logger.info(f"审核测试用例: {test_case.name} -> {test_case.status} (action={action})")

    return TestCaseResponse.model_validate(test_case)


@router.post("/batch-review")
def batch_review_test_cases(
    req: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_UPDATE)
):
    """
    批量审核测试用例

    功能要求（需求文档3.3.3）:
    - 批量审核通过/驳回
    - 支持 action: approve / reject
    """
    case_ids = req.get("case_ids", [])
    action = req.get("action", "approve")
    comment = req.get("comment")

    if not case_ids:
        raise HTTPException(status_code=400, detail="case_ids 不能为空")

    # 可审核状态集（与前端批量审核预检同源）：draft/pending_review/rejected 可审；
    # 已通过/已发布/终态（approved/published/deprecated/archived）跳过——
    # 全选批量审核时已审核记录会被带入，重复审核会刷新 reviewed_at 并误导计数
    _REVIEWABLE = {"draft", "pending_review", "rejected"}

    from datetime import datetime
    now = datetime.utcnow()
    updated = 0
    skipped_ids = []

    test_cases = db.query(TestCase).filter(TestCase.id.in_(case_ids)).all()
    for tc in test_cases:
        if (tc.status or "") not in _REVIEWABLE:
            skipped_ids.append(str(tc.id))
            continue
        if action == "approve":
            tc.status = TestCaseStatus.APPROVED.value
        elif action == "reject":
            tc.status = TestCaseStatus.REJECTED.value
        else:
            raise HTTPException(status_code=400, detail=f"不支持的操作: {action}")
        tc.reviewer_id = current_user["user"].id
        tc.review_comment = comment
        tc.reviewed_at = now
        updated += 1

    db.commit()

    logger.info(f"批量审核测试用例: {updated}/{len(case_ids)} 条 -> action={action}（跳过 {len(skipped_ids)} 条已审核/终态）")

    return {"success": True, "updated": updated, "total": len(case_ids),
            "skipped_count": len(skipped_ids), "skipped_ids": skipped_ids}


@router.get("/modules/{project_id}")
def get_modules(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取项目的模块列表"""
    modules = db.query(TestCase.module).filter(
        TestCase.project_id == project_id,
        TestCase.module.isnot(None)
    ).distinct().all()
    
    return {"modules": [m[0] for m in modules if m[0]]}


@router.post("/import", response_model=dict)
async def import_test_cases_from_file(
    version_id: int = Query(..., description="版本ID"),
    project_id: int = Query(..., description="项目ID"),
    file: UploadFile = File(..., description="CSV或Excel文件"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE),
):
    """
    从CSV/Excel直接导入功能测试用例（不走LLM，直接入库）

    支持格式: .csv / .xlsx / .xls
    表头映射（中英文均支持）:
      - 模块/module, 用例名称/name/title, 前置条件/precondition/preconditions
      - 测试步骤/steps/test_steps, 预期结果/expected_result/expected
      - 优先级/priority, 用例类型/case_type/type, 标签/tags
    导入后 source=imported_cases, status=published（已评审通过）
    """
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(400, f"不支持的文件格式: .{ext}，支持 .csv / .xlsx / .xls")

    content = await file.read()

    # ===== 解析文件为行列表 =====
    rows = []
    if ext == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(500, "服务器未安装 openpyxl，无法解析 Excel。请使用 CSV 格式。")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, h in enumerate(headers):
                if h:
                    row_dict[h] = str(row[i]) if row[i] is not None else ""
            if any(v for v in row_dict.values()):
                rows.append(row_dict)
        wb.close()

    if not rows:
        raise HTTPException(400, "文件中没有数据行")

    # ===== 列名映射 =====
    COLUMN_MAP = {
        "模块": "module", "module": "module",
        "用例名称": "name", "标题": "name", "name": "name", "title": "name",
        "前置条件": "preconditions", "precondition": "preconditions", "preconditions": "preconditions",
        "测试步骤": "test_steps_text", "步骤": "test_steps_text", "steps": "test_steps_text", "test_steps": "test_steps_text",
        "预期结果": "expected_result", "expected": "expected_result", "expected_result": "expected_result",
        "优先级": "priority", "priority": "priority",
        "用例类型": "case_type", "类型": "case_type", "type": "case_type", "case_type": "case_type",
        "标签": "tags_text", "tags": "tags_text",
    }

    imported = 0
    errors = []
    base_order = db.query(TestCase).filter(TestCase.version_id == version_id).count() * 10 + 10

    for i, row in enumerate(rows):
        # 映射列名
        mapped = {}
        for key, val in row.items():
            target = COLUMN_MAP.get(key.strip(), key.strip())
            if target not in mapped:
                mapped[target] = val.strip() if val else ""

        name = mapped.get("name", "")
        if not name:
            errors.append(f"第{i+2}行: 缺少用例名称")
            continue

        # 解析测试步骤（支持两种格式：JSON数组 或 换行分隔文本）
        steps_text = mapped.get("test_steps_text", "")
        steps = []
        if steps_text:
            try:
                parsed = json.loads(steps_text)
                if isinstance(parsed, list):
                    steps = parsed
            except Exception:
                # 按换行拆分，每行一个步骤
                for line in steps_text.split("\n"):
                    line = line.strip()
                    if line:
                        steps.append({"step": len(steps) + 1, "action": line, "expected": ""})

        # 解析标签
        tags_text = mapped.get("tags_text", "")
        tags = [t.strip() for t in tags_text.replace("；", ";").replace(",", ";").split(";") if t.strip()]

        try:
            tc = TestCase(
                project_id=project_id,
                version_id=version_id,
                module=mapped.get("module") or "未分类",
                name=name,
                description=mapped.get("description", ""),
                preconditions=mapped.get("preconditions", ""),
                test_steps=steps,
                expected_result=mapped.get("expected_result", ""),
                priority=mapped.get("priority", "P2") if mapped.get("priority") in ("P0","P1","P2","P3") else "P2",
                case_type=mapped.get("case_type", "functional"),
                status="published",       # 已评审通过的用例，直接发布
                tags=tags,
                generated_by="manual",    # 手动导入，不是AI生成
                sort_order=base_order + imported * 10,
                created_by=current_user.get("user").id if isinstance(current_user, dict) else 1,
            )
            db.add(tc)
            db.flush()
            # 方案B：新建用例逻辑=物理（logical_case_id=自身id）
            tc.logical_case_id = tc.id
            imported += 1
        except Exception as e:
            errors.append(f"第{i+2}行 '{name}': {str(e)}")

    db.commit()

    logger.info(f"用例导入完成: {imported} 条成功, {len(errors)} 条失败")
    return {
        "success": True,
        "imported": imported,
        "total": len(rows),
        "errors": errors[:10],
    }