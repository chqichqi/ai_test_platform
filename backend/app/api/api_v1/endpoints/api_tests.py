"""
API接口测试API端点
对应需求文档 3.5 API接口测试
"""

from typing import Optional, List, Dict, Any
from datetime import datetime
import json
import re
import io
import csv
import os

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
from bs4 import BeautifulSoup
from typing import Optional

from app.core.database import get_db
from app.core.models.api_test import ApiDefinition, ApiEndpoint, ApiTestCase, ApiTestExecution, ApiEnvironment, ApiTestVersion
from app.core.models.project import Project, Version
from app.core.schemas.api_test import (
    ImportSwaggerRequest, ApiDefinitionResponse,
    ApiEndpointResponse, ApiEndpointListResponse,
    ApiTestCaseCreate, ApiTestCaseUpdate, ApiTestCaseResponse, ApiTestCaseListResponse,
    ExecuteApiTestRequest, ApiTestExecutionResponse,
    GenerateApiTestCasesRequest, GenerateApiTestCasesResponse,
    ApiEnvironmentCreate, ApiEnvironmentResponse,
    SwaggerAutoGenerateRequest, SwaggerAutoGenerateResponse, GeneratedApiTestCase,
    ApiTestVersionCreate, ApiTestVersionResponse, ApiTestVersionListResponse,
    BatchDeleteRequest, BatchExecuteRequest, BatchExecuteResponse,
    SubmitReviewRequest, ReviewActionRequest, ReviewStatisticsResponse,
    ExportQueryParams, ReportRequest, ReportResponse,
    AuthConfigSchema, TokenExtractionSchema, TokenInjectionSchema,
    TestAuthRequest, TestAuthResponse, FileHashResponse,
)
from app.core.logger import logger
from app.core.middleware.permission_middleware import Permissions
# 非 JSON 原文 body 截断上限（与生成侧 api_flow_capture 同源——A3 修复 2026-08-25：
# 生成侧已截断，执行侧为手工录入超长 body 的防御；常量同源防漂移）
from app.core.services.api_flow_capture import _RAW_BODY_MAX_CHARS
from app.core.services.test_data_manager import TestDataManager
from app.core.services.test_data_plan import build_api_test_data_plan

router = APIRouter()


def normalize_url_colon(url: str) -> str:
    """转换URL中的中文冒号为英文冒号
    
    Args:
        url: 可能包含中文冒号的URL
        
    Returns:
        转换后的URL（中文冒号转为英文冒号）
    """
    if not url:
        return url
    
    # 中文冒号：英文冒号:
    return url.replace('：', ':')

# ===== 环境鉴权Token缓存 =====
import time as _time
_token_cache: Dict[str, tuple] = {}  # {cache_key: (token, expiry_timestamp)}  cache_key: env_{id} / project_{id}


def _get_cached_token(cache_key) -> Optional[str]:
    """获取缓存的token（未过期）"""
    if cache_key in _token_cache:
        token, expiry = _token_cache[cache_key]
        if _time.time() < expiry:
            return token
        del _token_cache[cache_key]
    return None


def _cache_token(cache_key, token: str, ttl: int = 3600) -> None:
    """缓存token"""
    _token_cache[cache_key] = (token, _time.time() + ttl)


def _get_project_api_auth(db, project_id) -> Optional[dict]:
    """读取项目级 API 鉴权（登录模块联动保存的 exploration_config.api_auth）。

    仅当已真实验证（verified=True，token 在浏览器登录时确实拿到过）才自动应用。
    """
    if not db or not project_id:
        return None
    try:
        from app.core.models.project_ext import ProjectSetting
        ps = db.query(ProjectSetting).filter(ProjectSetting.project_id == project_id).first()
        if not ps:
            return None
        auth = (ps.exploration_config or {}).get("api_auth") or {}
        if not auth.get("verified") or not auth.get("login_url"):
            return None
        return auth
    except Exception:
        return None


def _project_auth_to_env_config(db, project_id: int, project_auth: dict) -> dict:
    """项目级 api_auth → 执行链路 auth_config 结构。

    - 请求体占位符 {username}/{password} 在转换时替换为真实凭证（不落库，仅本次请求使用）
    - token_path/token_source → token_extraction；token_inject_* → token_injection
    """
    username = password = ""
    try:
        from app.core.models.project_ext import ProjectSetting
        ps = db.query(ProjectSetting).filter(ProjectSetting.project_id == project_id).first()
        web = (ps.exploration_config or {}).get("web", {}) if ps else {}
        username = web.get("username", "") or ""
        password = web.get("password", "") or ""
    except Exception:
        pass
    body = {}
    for k, v in (project_auth.get("request_body") or {}).items():
        if isinstance(v, str):
            body[k] = v.replace("{username}", username).replace("{password}", password)
        else:
            body[k] = v
    source = project_auth.get("token_source", "body")
    path = project_auth.get("token_path", "data.token") or "data.token"
    template = project_auth.get("token_inject_template") or "Bearer {token}"
    return {
        "enabled": True,
        "auth_type": "login_api",
        "login_url": project_auth.get("login_url", ""),
        "login_method": project_auth.get("login_method", "POST"),
        "login_body": body,
        "content_type": "application/json",
        "token_extraction": {
            "source": source,
            "json_path": path if source == "body" else None,
            "header_name": path if source == "header" else None,
        },
        "token_injection": {
            "header_name": project_auth.get("token_inject_name") or "Authorization",
            # "Bearer {token}" → "Bearer "（注入时再拼 token）
            "prefix": str(template).split("{token}")[0] or "",
        },
    }


async def _execute_auth_config(cache_key, base_url: str, auth_config: dict, source_label: str) -> dict:
    """按 auth_config 执行鉴权登录并提取 Token（环境级与项目级共用的执行器）。

    Returns: {auth_token, auth_header, api_keys, cookies, token_injection, logs}
    token_injection 透传注入参数（header_name/prefix），供用例执行时注入请求头。
    """
    auth_type = auth_config.get("auth_type", "bearer_token")
    token_cache_duration = auth_config.get("token_cache_duration", 3600)
    result = {"logs": []}

    # 检查缓存
    cached_token = _get_cached_token(cache_key)
    if cached_token:
        logger.info(f"使用缓存的鉴权token ({source_label}, key={cache_key})")
        result["auth_token"] = cached_token
        result["logs"].append("🔐 使用缓存的鉴权Token")
        return result

    try:
        # Bearer Token — 静态Token
        if auth_type == "bearer_token":
            inj = auth_config.get("token_injection", {})
            header_name = inj.get("header_name", "Authorization") if isinstance(inj, dict) else "Authorization"
            prefix = inj.get("prefix", "Bearer ") if isinstance(inj, dict) else "Bearer "
            token = auth_config.get("credentials", {}).get("token", "") if isinstance(auth_config.get("credentials"), dict) else ""
            if token:
                result["auth_token"] = token
                result["token_injection"] = {"header_name": header_name, "prefix": prefix}
                _cache_token(cache_key, token, token_cache_duration)
                result["logs"].append(f"🔐 Bearer Token鉴权: {header_name}")
            return result

        # Basic Auth
        if auth_type == "basic_auth":
            import base64
            creds = auth_config.get("credentials", {}) or {}
            username = creds.get("username", "")
            password = creds.get("password", "")
            encoded = base64.b64encode(f"{username}:{password}".encode()).decode()
            result["auth_header"] = f"Basic {encoded}"
            result["logs"].append("🔐 Basic Auth鉴权")
            return result

        # API Key
        if auth_type == "api_key":
            inj = auth_config.get("token_injection", {}) or {}
            header_name = inj.get("header_name", "X-API-Key") if isinstance(inj, dict) else "X-API-Key"
            creds = auth_config.get("credentials", {}) or {}
            api_key = creds.get("api_key", "")
            result["api_keys"] = {header_name: api_key}
            result["logs"].append(f"🔐 API Key鉴权: {header_name}")
            return result

        # Cookie
        if auth_type == "cookie":
            creds = auth_config.get("credentials", {}) or {}
            cookie_name = creds.get("cookie_name", "")
            cookie_value = creds.get("cookie_value", "")
            if cookie_name:
                result["cookies"] = {cookie_name: cookie_value}
                result["logs"].append(f"🔐 Cookie鉴权: {cookie_name}")
            return result

        # Login API (bearer_token with login) / OAuth2
        login_url = auth_config.get("login_url") or auth_config.get("token_url")
        if not login_url:
            return result

        login_method = auth_config.get("login_method", "POST")
        login_headers = dict(auth_config.get("login_headers", {}) or {})
        login_body = auth_config.get("login_body") or auth_config.get("credentials") or {}
        content_type = auth_config.get("content_type", "application/json")

        full_url = base_url.rstrip("/") + "/" + login_url.lstrip("/")
        result["logs"].append(f"🔐 正在鉴权登录: {login_method} {full_url}")

        async with httpx.AsyncClient(timeout=30) as client:
            req_kwargs = {
                "method": login_method,
                "url": full_url,
                "headers": login_headers if login_headers else None,
            }

            if "application/x-www-form-urlencoded" in content_type:
                req_kwargs["data"] = login_body
                if "Content-Type" not in login_headers:
                    login_headers["Content-Type"] = "application/x-www-form-urlencoded"
                    req_kwargs["headers"] = login_headers
            else:
                req_kwargs["json"] = login_body
                if "Content-Type" not in login_headers:
                    login_headers["Content-Type"] = "application/json"
                    req_kwargs["headers"] = login_headers

            response = await client.request(**req_kwargs)

            if response.status_code in [200, 201]:
                try:
                    response_body = response.json()
                except Exception:
                    response_body = {}

                # Token提取
                extraction = auth_config.get("token_extraction") or {}
                if isinstance(extraction, str):
                    try:
                        extraction = json.loads(extraction)
                    except Exception:
                        extraction = {}
                source = extraction.get("source", "body") if isinstance(extraction, dict) else "body"
                json_path = extraction.get("json_path", "data.token") if isinstance(extraction, dict) else "data.token"
                header_name = extraction.get("header_name") if isinstance(extraction, dict) else None
                cookie_name = extraction.get("cookie_name") if isinstance(extraction, dict) else None

                token = None
                if source == "body" and json_path:
                    token = _extract_value_from_dict(response_body, json_path)
                elif source == "header" and header_name:
                    token = response.headers.get(header_name)
                    if token and token.startswith("Bearer "):
                        token = token[7:]
                elif source == "cookie" and cookie_name:
                    token = response.cookies.get(cookie_name)

                if token:
                    result["auth_token"] = token
                    # 注入参数透传（Bearer {token} 模板 → prefix + token；默认 Authorization: Bearer x）
                    inj = auth_config.get("token_injection") or {}
                    if isinstance(inj, str):
                        try:
                            inj = json.loads(inj)
                        except Exception:
                            inj = {}
                    result["token_injection"] = {
                        "header_name": (inj or {}).get("header_name", "Authorization"),
                        "prefix": (inj or {}).get("prefix", "Bearer "),
                    }
                    _cache_token(cache_key, token, token_cache_duration)
                    result["logs"].append(f"✅ 鉴权成功，Token已缓存({token_cache_duration}s)")

                    # Cookie传递
                    if response.cookies:
                        result["cookies"] = dict(response.cookies)
                else:
                    result["logs"].append("⚠️ 鉴权登录成功但未能提取Token")
                    # 兜底：尝试常见token字段
                    for field in ["token", "access_token", "data.token", "data.access_token"]:
                        token = _extract_value_from_dict(response_body, field)
                        if token:
                            result["auth_token"] = token
                            _cache_token(cache_key, token, token_cache_duration)
                            result["logs"].append(f"✅ 从字段'{field}'自动提取Token成功")
                            break
            else:
                result["logs"].append(f"❌ 鉴权登录失败: HTTP {response.status_code}")
                result["auth_error"] = f"HTTP {response.status_code}: {response.text[:200]}"

    except Exception as e:
        result["logs"].append(f"❌ 鉴权异常: {str(e)}")
        result["auth_error"] = str(e)
        logger.error(f"{source_label}鉴权失败 (key={cache_key}): {str(e)}")

    return result


async def _execute_env_auth(env, base_url: str, db=None, project_id: int = None) -> dict:
    """执行鉴权登录（环境级优先，缺失时回退项目级 API 鉴权——登录模块联动保存的）。

    Args:
        env: ApiEnvironment对象（可能为 None：项目未建默认环境）
        base_url: 基础URL
        db: 数据库会话（回退项目级 api_auth 时需要）
        project_id: 项目ID（env 为 None 时回退项目级鉴权必须显式传入）

    Returns:
        {auth_token, auth_header, api_keys, cookies, token_injection, logs, auth_source}
        auth_source: env_config（环境配置）| project_linked（项目联动）
    """
    # 1. 环境级显式配置优先（用户精细配置 > 项目联动默认）
    if env is not None and env.auth_config:
        auth_config = env.auth_config
        if isinstance(auth_config, str):
            try:
                auth_config = json.loads(auth_config)
            except Exception:
                auth_config = None
        if auth_config and auth_config.get("enabled"):
            result = await _execute_auth_config(f"env_{env.id}", base_url, auth_config, "环境")
            result["auth_source"] = "env_config"
            return result

    # 2. 回退：项目级 API 鉴权（已验证的登录模块联动配置）——Swagger 导入生成的用例自动可用
    if db is not None:
        pid = project_id if project_id is not None else (env.project_id if env is not None else None)
        project_auth = _get_project_api_auth(db, pid)
        if project_auth:
            auth_config = _project_auth_to_env_config(db, pid, project_auth)
            # 缓存键含 verified_at 指纹：配置变更（重新联动/手动改配置会重置 verified_at）→ 旧缓存自动失效
            cache_key = f"project_{pid}_{(project_auth.get('verified_at') or 'na')}"
            result = await _execute_auth_config(cache_key, base_url, auth_config, "项目联动")
            result["auth_source"] = "project_linked"
            return result

    return {}


def _parse_swagger_html(html_content: str) -> Optional[Dict[str, Any]]:
    """解析Swagger UI HTML页面，提取API信息"""
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        
        swagger_data = {
            "openapi": "3.0.0",
            "info": {
                "title": "API文档",
                "version": "1.0.0"
            },
            "paths": {}
        }
        
        # 尝试从script中提取spec
        scripts = soup.find_all('script')
        for script in scripts:
            if script.string:
                content = script.string
                if 'spec' in content or 'swagger' in content.lower():
                    json_match = re.search(r'"spec"\s*:\s*(\{[\s\S]*?\})', content)
                    if json_match:
                        try:
                            spec_json = json.loads(json_match.group(1))
                            if "paths" in spec_json:
                                return spec_json
                        except:
                            pass
                    
                    json_match = re.search(r'(\{[\s\S]*"paths"[\s\S]*\})', content)
                    if json_match:
                        try:
                            return json.loads(json_match.group(1))
                        except:
                            pass
        
        # 从DOM元素中提取
        path_sections = soup.find_all(['div', 'section'], class_=re.compile(r'(path|endpoint|operation)', re.I))
        
        for section in path_sections:
            path_elem = section.find(['span', 'div', 'a'], class_=re.compile(r'(path|route)', re.I))
            method_elem = section.find(['span', 'div'], class_=re.compile(r'(method|verb|get|post|put|delete|patch)', re.I))
            
            if path_elem and method_elem:
                path_text = path_elem.get_text(strip=True)
                method_text = method_elem.get_text(strip=True).upper()
                
                if method_text in ["GET", "POST", "PUT", "DELETE", "PATCH"]:
                    summary_elem = section.find(['span', 'div', 'p'], class_=re.compile(r'(summary|description)', re.I))
                    summary = summary_elem.get_text(strip=True) if summary_elem else ""
                    
                    if path_text not in swagger_data["paths"]:
                        swagger_data["paths"][path_text] = {}
                    
                    swagger_data["paths"][path_text][method_text.lower()] = {
                        "summary": summary,
                        "responses": {"200": {"description": "成功"}}
                    }
        
        title_elem = soup.find(['title', 'h1', 'h2'])
        if title_elem:
            swagger_data["info"]["title"] = title_elem.get_text(strip=True)
        
        if swagger_data["paths"]:
            logger.info(f"Parsed {len(swagger_data['paths'])} paths from HTML")
            return swagger_data
        
        return None
        
    except Exception as e:
        logger.error(f"Failed to parse Swagger HTML: {str(e)}")
        return None


@router.post("/import", response_model=ApiDefinitionResponse, status_code=status.HTTP_201_CREATED)
async def import_swagger(
    request: ImportSwaggerRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """
    导入Swagger/OpenAPI文档
    
    支持URL导入和文件上传，支持JSON、YAML和Swagger UI页面（如/docs）
    """
    project = db.query(Project).filter(Project.id == request.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail=f"项目ID {request.project_id} 不存在")
    
    swagger_content = None
    base_url = None
    
    if request.source_type == "url" and request.source_url:
        # 转换中文冒号为英文冒号
        source_url = normalize_url_colon(request.source_url)
        
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.get(source_url)
            response.raise_for_status()
            
            content_type = response.headers.get("content-type", "").lower()
            content = response.text
            
            # 尝试解析JSON
            if "json" in content_type or content.strip().startswith("{") or content.strip().startswith("["):
                try:
                    swagger_content = response.json()
                except:
                    pass
            
            # 尝试解析YAML
            if not swagger_content and ("yaml" in content_type or "yml" in request.source_url.lower()):
                try:
                    import yaml
                    swagger_content = yaml.safe_load(content)
                except:
                    pass
            
            # 处理Swagger UI页面（如/docs）
            if not swagger_content and ("/docs" in request.source_url or request.source_url.endswith("/docs") or request.source_url.endswith("/docs/")):
                # 尝试获取openapi.json
                json_url = request.source_url.rstrip("/")
                if json_url.endswith("/docs"):
                    json_url = json_url.replace("/docs", "/openapi.json")
                else:
                    json_url = request.source_url.replace("/docs", "/openapi.json")
                
                logger.info(f"Detected Swagger UI page, trying JSON URL: {json_url}")
                
                try:
                    json_response = await client.get(json_url)
                    json_response.raise_for_status()
                    swagger_content = json_response.json()
                    logger.info(f"Successfully fetched OpenAPI JSON from {json_url}")
                except Exception as e:
                    logger.warning(f"Failed to fetch JSON from {json_url}: {e}")
            
            # 如果是HTML页面，尝试解析
            if not swagger_content and ("html" in content_type or content.strip().startswith("<") or "<!doctype" in content.lower()):
                logger.info("Detected HTML page, trying to extract API info")
                swagger_content = _parse_swagger_html(content)
            
            # 最后尝试直接解析
            if not swagger_content:
                try:
                    swagger_content = response.json()
                except:
                    try:
                        import yaml
                        swagger_content = yaml.safe_load(content)
                    except:
                        pass
    else:
        raise HTTPException(status_code=400, detail="请提供Swagger文档URL")
    
    if not swagger_content:
        raise HTTPException(status_code=400, detail="无法获取Swagger文档内容，请确保URL指向正确的OpenAPI JSON/YAML文件，或使用 /openapi.json 而非 /docs")
    
    openapi_version = swagger_content.get("openapi") or swagger_content.get("swagger", "2.0")
    
    if "servers" in swagger_content and swagger_content["servers"]:
        base_url = swagger_content["servers"][0].get("url")
    elif "host" in swagger_content:
        base_url = f"http://{swagger_content['host']}{swagger_content.get('basePath', '')}"
    
    definition = ApiDefinition(
        project_id=request.project_id,
        name=request.name or swagger_content.get("info", {}).get("title", "API文档"),
        source_type=request.source_type,
        source_url=request.source_url,
        content=swagger_content,
        version=openapi_version,
        base_url=base_url,
        description=swagger_content.get("info", {}).get("description"),
        imported_at=datetime.utcnow()
    )
    
    db.add(definition)
    db.flush()
    
    endpoints = _parse_swagger_endpoints(definition.id, swagger_content)
    for endpoint in endpoints:
        db.add(endpoint)
    
    db.commit()
    db.refresh(definition)
    
    logger.info(f"导入Swagger文档: {definition.name}, 接口数: {len(endpoints)}")
    
    return ApiDefinitionResponse.model_validate(definition)


@router.post("/import/file", response_model=ApiDefinitionResponse, status_code=status.HTTP_201_CREATED)
async def import_swagger_file(
    project_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """上传Swagger/OpenAPI JSON/YAML文件"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail=f"项目ID {project_id} 不存在")
    
    content = await file.read()
    
    try:
        if file.filename and file.filename.endswith(('.yaml', '.yml')):
            import yaml
            swagger_content = yaml.safe_load(content.decode('utf-8'))
        else:
            swagger_content = json.loads(content.decode('utf-8'))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析文件失败: {str(e)}")
    
    openapi_version = swagger_content.get("openapi") or swagger_content.get("swagger", "2.0")
    base_url = None
    
    if "servers" in swagger_content and swagger_content["servers"]:
        base_url = swagger_content["servers"][0].get("url")
    elif "host" in swagger_content:
        base_url = f"http://{swagger_content['host']}{swagger_content.get('basePath', '')}"
    
    definition = ApiDefinition(
        project_id=project_id,
        name=swagger_content.get("info", {}).get("title", file.filename or "API文档"),
        source_type="file",
        content=swagger_content,
        version=openapi_version,
        base_url=base_url,
        description=swagger_content.get("info", {}).get("description"),
        imported_at=datetime.utcnow()
    )
    
    db.add(definition)
    db.flush()
    
    endpoints = _parse_swagger_endpoints(definition.id, swagger_content)
    for endpoint in endpoints:
        db.add(endpoint)
    
    db.commit()
    db.refresh(definition)
    
    logger.info(f"上传Swagger文档: {definition.name}")
    
    return ApiDefinitionResponse.model_validate(definition)


def _parse_swagger_endpoints(definition_id: int, swagger: dict) -> List[ApiEndpoint]:
    """解析Swagger文档中的接口"""
    endpoints = []
    paths = swagger.get("paths", {})
    
    for path, methods in paths.items():
        for method, spec in methods.items():
            if method.upper() not in ["GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS"]:
                continue
            
            tags = spec.get("tags", [])
            tag = tags[0] if tags else None
            
            endpoint = ApiEndpoint(
                definition_id=definition_id,
                path=path,
                method=method.upper(),
                tag=tag,
                summary=spec.get("summary"),
                description=spec.get("description"),
                parameters=spec.get("parameters"),
                request_body=spec.get("requestBody") or spec.get("body"),
                responses=spec.get("responses"),
                security=spec.get("security"),
                deprecated=spec.get("deprecated", False)
            )
            endpoints.append(endpoint)
    
    return endpoints


@router.get("/definitions/{project_id}", response_model=List[ApiDefinitionResponse])
def list_definitions(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取项目的API定义列表"""
    definitions = db.query(ApiDefinition).filter(
        ApiDefinition.project_id == project_id
    ).order_by(ApiDefinition.imported_at.desc()).all()
    
    return [ApiDefinitionResponse.model_validate(d) for d in definitions]


@router.get("/endpoints/{definition_id}", response_model=ApiEndpointListResponse)
def list_endpoints(
    definition_id: int,
    tag: Optional[str] = Query(None),
    method: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取API定义的接口列表"""
    query = db.query(ApiEndpoint).filter(ApiEndpoint.definition_id == definition_id)
    
    if tag:
        query = query.filter(ApiEndpoint.tag == tag)
    if method:
        query = query.filter(ApiEndpoint.method == method.upper())
    
    total = query.count()
    endpoints = query.offset((page - 1) * page_size).limit(page_size).all()
    
    return ApiEndpointListResponse(
        items=[ApiEndpointResponse.model_validate(e) for e in endpoints],
        total=total,
        page=page,
        page_size=page_size
    )


@router.get("/tags/{definition_id}")
def get_tags(
    definition_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取API定义的标签列表"""
    tags = db.query(ApiEndpoint.tag).filter(
        ApiEndpoint.definition_id == definition_id,
        ApiEndpoint.tag.isnot(None)
    ).distinct().all()
    
    return {"tags": [t[0] for t in tags if t[0]]}


@router.post("/cases", response_model=ApiTestCaseResponse, status_code=status.HTTP_201_CREATED)
def create_test_case(
    case_in: ApiTestCaseCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """创建API测试用例"""
    test_case = ApiTestCase(
        project_id=case_in.project_id,
        endpoint_id=case_in.endpoint_id,
        version_id=case_in.version_id,
        name=case_in.name,
        description=case_in.description,
        method=case_in.method,
        path=case_in.path,
        base_url=case_in.base_url,
        headers=case_in.headers,
        query_params=case_in.query_params,
        path_params=case_in.path_params,
        request_body=case_in.request_body,
        test_data=case_in.test_data,
        expected_status=case_in.expected_status,
        expected_headers=case_in.expected_headers,
        expected_body=case_in.expected_body,
        assert_rules=[r.model_dump() for r in case_in.assert_rules] if case_in.assert_rules else None,
        case_type=case_in.case_type,
        priority=case_in.priority,
        tags=case_in.tags,
        depends_on=case_in.depends_on,
        variable_extractions=[v.model_dump() for v in case_in.variable_extractions] if case_in.variable_extractions else None,
        generated_by="manual",
        created_by=current_user["user"].id
    )
    
    db.add(test_case)
    db.commit()
    db.refresh(test_case)
    
    logger.info(f"创建API测试用例: {test_case.name}")
    
    return ApiTestCaseResponse.model_validate(test_case)


@router.get("/cases", response_model=ApiTestCaseListResponse)
def list_test_cases(
    project_id: int = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    case_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """API测试用例列表"""
    query = db.query(ApiTestCase).filter(ApiTestCase.project_id == project_id)
    
    if case_type:
        query = query.filter(ApiTestCase.case_type == case_type)
    if priority:
        query = query.filter(ApiTestCase.priority == priority)
    if search:
        pattern = f"%{search}%"
        query = query.filter(ApiTestCase.name.ilike(pattern))
    
    total = query.count()
    total_pages = (total + page_size - 1) // page_size
    
    cases = query.order_by(ApiTestCase.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return ApiTestCaseListResponse(
        items=[ApiTestCaseResponse.model_validate(c) for c in cases],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.post("/cases/generate", response_model=GenerateApiTestCasesResponse)
def generate_test_cases(
    request: GenerateApiTestCasesRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """
    AI生成API测试用例
    
    根据接口定义自动生成测试用例
    """
    endpoint = db.query(ApiEndpoint).filter(ApiEndpoint.id == request.endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail=f"接口ID {request.endpoint_id} 不存在")
    
    definition = db.query(ApiDefinition).filter(ApiDefinition.id == endpoint.definition_id).first()
    
    generated_cases = []
    
    if request.include_normal:
        case = _generate_normal_case(endpoint, definition, current_user["user"].id)
        db.add(case)
        generated_cases.append(case)
    
    if request.include_error:
        error_cases = _generate_error_cases(endpoint, definition, current_user["user"].id)
        for case in error_cases:
            db.add(case)
            generated_cases.append(case)
    
    if request.include_boundary:
        boundary_cases = _generate_boundary_cases(endpoint, definition, current_user["user"].id)
        for case in boundary_cases:
            db.add(case)
            generated_cases.append(case)
    
    db.commit()
    
    for case in generated_cases:
        db.refresh(case)
    
    logger.info(f"AI生成API测试用例: {len(generated_cases)} 条")
    
    return GenerateApiTestCasesResponse(
        generated_count=len(generated_cases),
        test_cases=[ApiTestCaseResponse.model_validate(c) for c in generated_cases]
    )


def _generate_normal_case(endpoint: ApiEndpoint, definition: ApiDefinition, user_id: int) -> ApiTestCase:
    """统一 OpenAPI 生成器：单接口正常场景。"""
    from app.core.services.openapi_test_generator import OpenApiTestGenerator
    ep = {"method": endpoint.method, "path": endpoint.path, "summary": endpoint.summary or "",
          "description": endpoint.description or "", "parameters": endpoint.parameters or [],
          "request_body": endpoint.request_body or {}, "responses": endpoint.responses or {},
          "security": endpoint.security, "requires_auth": bool(endpoint.security)}
    gen = OpenApiTestGenerator()
    data = gen.generate_test_cases(ep, include_normal=True, include_error=False, include_boundary=False,
                                   include_auth=False, max_cases=1)[0]
    return ApiTestCase(
        project_id=definition.project_id, endpoint_id=endpoint.id,
        name=data["name"], description=data.get("description", ""), method=endpoint.method, path=endpoint.path,
        base_url=definition.base_url, query_params=data.get("query_params") or None,
        path_params=data.get("path_params") or None, request_body=data.get("request_body") or None,
        headers=data.get("headers") or None,
        test_data=build_api_test_data_plan(data.get("query_params"), data.get("path_params"), data.get("request_body"), data.get("headers"),
                                           metadata={"source":"unified_openapi_generator","variant":"normal"}),
        expected_status=data.get("expected_status", 200), assert_rules=data.get("assert_rules", []),
        test_steps=data.get("test_steps", []), expected_result=data.get("expected_result", ""),
        preconditions=data.get("preconditions", ""), case_type="normal", priority="P1",
        status="draft", generated_by="ai", created_by=user_id
    )


def _generate_error_cases(endpoint: ApiEndpoint, definition: ApiDefinition, user_id: int) -> List[ApiTestCase]:
    """统一 OpenAPI 生成器：只生成接口实际具备意义的异常变体。"""
    from app.core.services.openapi_test_generator import OpenApiTestGenerator
    ep = {"method": endpoint.method, "path": endpoint.path, "summary": endpoint.summary or "",
          "description": endpoint.description or "", "parameters": endpoint.parameters or [],
          "request_body": endpoint.request_body or {}, "responses": endpoint.responses or {},
          "security": endpoint.security, "requires_auth": bool(endpoint.security)}
    gen = OpenApiTestGenerator()
    data_list = gen.generate_test_cases(ep, include_normal=False, include_error=True, include_boundary=False,
                                        include_auth=True, max_cases=6)
    out = []
    for data in data_list:
        test_data = build_api_test_data_plan(
            data.get("query_params"), data.get("path_params"), data.get("request_body"), data.get("headers"),
            mutation_key=data.get("mutation_key", ""), mutation=data.get("mutation", ""),
            metadata={"source":"unified_openapi_generator", "variant":data.get("case_type","error")}
        )
        out.append(ApiTestCase(
            project_id=definition.project_id, endpoint_id=endpoint.id, name=data["name"],
            description=data.get("description", ""), method=endpoint.method, path=endpoint.path, base_url=definition.base_url,
            headers=data.get("headers") or None, query_params=data.get("query_params") or None,
            path_params=data.get("path_params") or None, request_body=data.get("request_body") or None,
            test_data=test_data, expected_status=data.get("expected_status"), assert_rules=data.get("assert_rules", []),
            test_steps=data.get("test_steps", []), expected_result=data.get("expected_result", ""),
            preconditions=data.get("preconditions", ""), case_type="error", priority="P2", status="draft",
            tags=["api_generated", "unified"], generated_by="ai", created_by=user_id
        ))
    return out


def _generate_boundary_cases(endpoint: ApiEndpoint, definition: ApiDefinition, user_id: int) -> List[ApiTestCase]:
    """统一 OpenAPI 生成器：边界场景。"""
    from app.core.services.openapi_test_generator import OpenApiTestGenerator
    ep = {"method": endpoint.method, "path": endpoint.path, "summary": endpoint.summary or "",
          "description": endpoint.description or "", "parameters": endpoint.parameters or [],
          "request_body": endpoint.request_body or {}, "responses": endpoint.responses or {},
          "security": endpoint.security, "requires_auth": bool(endpoint.security)}
    gen = OpenApiTestGenerator()
    data_list = gen.generate_boundary_cases(ep, max_cases=4)
    out=[]
    for data in data_list:
        out.append(ApiTestCase(
            project_id=definition.project_id, endpoint_id=endpoint.id, name=data["name"],
            description=data.get("description", ""), method=endpoint.method, path=endpoint.path, base_url=definition.base_url,
            query_params=data.get("query_params") or None, path_params=data.get("path_params") or None,
            request_body=data.get("request_body") or None, headers=data.get("headers") or None,
            test_data=build_api_test_data_plan(data.get("query_params"), data.get("path_params"), data.get("request_body"), data.get("headers"),
                                               metadata={"source":"unified_openapi_generator", "variant":"boundary"}),
            expected_status=data.get("expected_status"), assert_rules=data.get("assert_rules", []),
            test_steps=data.get("test_steps", []), expected_result=data.get("expected_result", ""),
            preconditions=data.get("preconditions", ""), case_type="boundary", priority="P3", status="draft",
            tags=["api_generated", "boundary", "unified"], generated_by="ai", created_by=user_id
        ))
    return out


async def _execute_precondition_cases(
    depends_on: list,
    base_url: str,
    db: Session,
    current_user: dict,
    credentials: Optional[HTTPAuthorizationCredentials],
    env_auth_vars: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """执行前置用例并提取变量

    env_auth_vars：本用例执行链路已提取的实时鉴权（execute_test 先行 _execute_env_auth）。
    探索生成用例被引用为前置时 headers 含 {{auth_token}} 占位符（不落明文 token），
    用实时 token 替换；无可用 token 则移除占位符头（避免发出字面占位符）。
    """
    extracted_vars = {}

    if not depends_on:
        return extracted_vars

    for case_id in depends_on:
        try:
            case_id_int = int(case_id)
            pre_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id_int).first()

            if not pre_case:
                logger.warning(f"Precondition case {case_id} not found")
                continue

            # 构建请求（F24 修复 2026-08-25：前置用例 base_url 优先于链路兜底——
            # 探索生成的用例自带捕获的真实 origin，忽略会导致跨域/404；与主链路
            # 「用例 base_url 优先」语义 M2 同源；rstrip('/') 防双斜杠拼接）
            _pre_base = (pre_case.base_url or base_url or "").rstrip("/")
            pre_url = normalize_url_colon(f"{_pre_base}{pre_case.path or '/'}")

            pre_headers = {}
            if pre_case.headers:
                if isinstance(pre_case.headers, dict):
                    pre_headers = dict(pre_case.headers)
            # 探索生成用例可能被引用为前置用例：headers 含 {{auth_token}} 占位符
            # （不落明文 token）。用本链路已提取的实时 token 替换；无可用 token 则
            # 移除占位符头（避免发出字面占位符）。
            if env_auth_vars and env_auth_vars.get("auth_token"):
                _pre_tok = env_auth_vars["auth_token"]
                _pre_inj = env_auth_vars.get("token_injection") or {}
                if isinstance(_pre_inj, str):
                    try:
                        _pre_inj = json.loads(_pre_inj)
                    except Exception:
                        _pre_inj = {}
                _pre_prefix = (_pre_inj.get("prefix") if isinstance(_pre_inj, dict) else None) or "Bearer "
                for _hk in list(pre_headers.keys()):
                    _hv = pre_headers[_hk]
                    if isinstance(_hv, str) and "{{auth_token}}" in _hv:
                        pre_headers[_hk] = _hv.replace("{{auth_token}}", f"{_pre_prefix}{_pre_tok}")
                        logger.info(f"前置用例鉴权占位符已替换: {_hk}")
            for _hk in list(pre_headers.keys()):
                if isinstance(pre_headers[_hk], str) and "{{auth_token}}" in pre_headers[_hk]:
                    pre_headers.pop(_hk, None)
                    logger.warning(f"前置用例鉴权占位符无 token 可替换，已移除请求头: {_hk}")

            pre_body = None
            if pre_case.request_body:
                if isinstance(pre_case.request_body, dict):
                    pre_body = dict(pre_case.request_body)
                elif isinstance(pre_case.request_body, str):
                    # A3 补全（2026-08-25）：前置引用路径与单条/批量主路径同构——
                    # 字符串 body 原样保留、发送时走 data=（json= 传字符串会被再编码）
                    pre_body = pre_case.request_body[:_RAW_BODY_MAX_CHARS]

            logger.info(f"Executing precondition case: {pre_case.method} {pre_url}")

            async with httpx.AsyncClient(timeout=30) as client:
                response = await client.request(
                    method=str(pre_case.method or "POST"),
                    url=pre_url,
                    headers=pre_headers if pre_headers else None,
                    json=pre_body if isinstance(pre_body, dict) and pre_body else None,
                    data=pre_body if isinstance(pre_body, str) else None
                )
                
                if response.status_code in [200, 201]:
                    try:
                        response_body = response.json()
                        
                        # 根据variable_extractions配置提取变量
                        if pre_case.variable_extractions:
                            for extraction in pre_case.variable_extractions:
                                var_name = extraction.get("name")
                                source = extraction.get("source", "response_body")
                                path = extraction.get("path", "")
                                
                                if source == "response_body" and path:
                                    # 支持嵌套路径提取 (如 "data.token", "token")
                                    value = _extract_value_from_dict(response_body, path)
                                    if value:
                                        extracted_vars[var_name] = value
                                        logger.info(f"Extracted variable: {var_name} = {value[:20]}...")
                        
                        # 如果没有配置variable_extractions，尝试自动提取常见token字段
                        if not pre_case.variable_extractions or not extracted_vars.get("auth_token"):
                            token_fields = ["token", "access_token", "auth_token", "data.token", "data.access_token"]
                            for field in token_fields:
                                value = _extract_value_from_dict(response_body, field)
                                if value:
                                    extracted_vars["auth_token"] = value
                                    logger.info(f"Auto-extracted auth_token from field: {field}")
                                    break
                        
                    except Exception as e:
                        logger.warning(f"Failed to parse precondition response: {str(e)}")
                else:
                    logger.warning(f"Precondition case failed with status: {response.status_code}")
                    
        except Exception as e:
            logger.error(f"Failed to execute precondition case {case_id}: {str(e)}")
    
    return extracted_vars


def _extract_value_from_dict(data: Dict[str, Any], path: str) -> Optional[Any]:
    """从字典中按路径提取值"""
    if not data or not path:
        return None
    
    keys = path.split(".")
    value = data
    
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
        else:
            return None
        
        if value is None:
            return None
    
    return value


def _build_dependency_graph(cases: List[ApiTestCase]) -> Dict[int, List[int]]:
    """构建用例依赖图
    
    Returns:
        {case_id: [依赖的case_id列表]}
    """
    graph = {}
    case_ids = {case.id for case in cases}
    
    for case in cases:
        depends_on = case.depends_on or []
        if isinstance(depends_on, list):
            graph[case.id] = [int(d) for d in depends_on if int(d) in case_ids]
        else:
            graph[case.id] = []
    
    return graph


def _detect_circular_dependency(graph: Dict[int, List[int]]) -> Optional[List[int]]:
    """检测循环依赖
    
    Returns:
        如果存在循环依赖，返回循环路径；否则返回None
    """
    visited = set()
    rec_stack = set()
    
    def dfs(node: int, path: List[int]) -> Optional[List[int]]:
        visited.add(node)
        rec_stack.add(node)
        path.append(node)
        
        for neighbor in graph.get(node, []):
            if neighbor not in visited:
                result = dfs(neighbor, path)
                if result:
                    return result
            elif neighbor in rec_stack:
                cycle_start_idx = path.index(neighbor)
                return path[cycle_start_idx:] + [neighbor]
        
        path.pop()
        rec_stack.remove(node)
        return None
    
    for node in graph:
        if node not in visited:
            cycle = dfs(node, [])
            if cycle:
                return cycle
    
    return None


def _topological_sort_cases(
    cases: List[ApiTestCase],
    all_cases_dict: Dict[int, ApiTestCase]
) -> List[ApiTestCase]:
    """拓扑排序用例，确保依赖的用例先执行
    
    Args:
        cases: 用户选择的用例列表
        all_cases_dict: 所有用例的字典（包括可能需要的前置用例）
    
    Returns:
        排序后的用例列表（包含需要的前置用例）
    """
    selected_ids = {case.id for case in cases}
    
    all_needed_ids = set(selected_ids)
    pending_check = list(selected_ids)
    
    while pending_check:
        case_id = pending_check.pop()
        case = all_cases_dict.get(case_id)
        if case and case.depends_on:
            depends_on = case.depends_on if isinstance(case.depends_on, list) else []
            for dep_id in depends_on:
                dep_id_int = int(dep_id)
                if dep_id_int not in all_needed_ids and dep_id_int in all_cases_dict:
                    all_needed_ids.add(dep_id_int)
                    pending_check.append(dep_id_int)
    
    graph = {}
    for case_id in all_needed_ids:
        case = all_cases_dict.get(case_id)
        if case:
            depends_on = case.depends_on if isinstance(case.depends_on, list) else []
            graph[case_id] = [int(d) for d in depends_on if int(d) in all_needed_ids]
        else:
            graph[case_id] = []
    
    cycle = _detect_circular_dependency(graph)
    if cycle:
        logger.warning(f"检测到循环依赖: {cycle}")
        for node in cycle:
            if node in graph:
                graph[node] = []
    
    # 入度 = 自己的依赖数（不是被依赖数）
    # 没有依赖的用例入度为0，应该最先执行
    in_degree = {}
    for node in all_needed_ids:
        # 入度 = 该节点依赖的其他节点数量
        dependencies = graph.get(node, [])
        in_degree[node] = len([d for d in dependencies if d in all_needed_ids])
    
    # 入度为0的节点（没有依赖）先执行
    queue = [node for node in all_needed_ids if in_degree[node] == 0]
    sorted_ids = []
    
    logger.info(f"拓扑排序: 总节点数={len(all_needed_ids)}, 入度为0的节点={queue}")
    
    while queue:
        # 取出入度为0的节点（没有依赖，可以执行）
        node = queue.pop(0)
        sorted_ids.append(node)
        
        # 执行node后，检查哪些依赖node的节点可以执行了
        # 需要找所有 graph[other] 中包含 node 的节点
        for other_node in all_needed_ids:
            if other_node != node and other_node not in sorted_ids:
                # 如果 other_node 依赖 node，执行node后other_node的入度减少
                if node in graph.get(other_node, []):
                    in_degree[other_node] -= 1
                    if in_degree[other_node] == 0:
                        queue.append(other_node)
    
    if len(sorted_ids) != len(all_needed_ids):
        remaining = [id for id in all_needed_ids if id not in sorted_ids]
        logger.warning(f"拓扑排序未完成，剩余节点: {remaining}")
        sorted_ids.extend(remaining)
    
    sorted_cases = [all_cases_dict[id] for id in sorted_ids if id in all_cases_dict]
    
    return sorted_cases, selected_ids


def _render_runtime_structure(obj, values: Dict[str, Any]):
    """递归替换 ${var}/{{var}}，同时支持 API 前置用例提取变量。"""
    if isinstance(obj, dict):
        return {k: _render_runtime_structure(v, values) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_render_runtime_structure(v, values) for v in obj]
    if isinstance(obj, str):
        text = obj
        for key, value in (values or {}).items():
            text = text.replace("${" + str(key) + "}", str(value))
            text = text.replace("{{" + str(key) + "}}", str(value))
        return text
    return obj


def _resolve_api_runtime_request(test_case):
    """统一实例化 API TestDataPlan。

    对普通历史用例允许兼容回退；对带 mutation 的异常/边界用例绝不回退到
    原始请求，否则“参数类型错误”会悄悄变成正常参数，造成严重假阴性。
    """
    try:
        manager = TestDataManager()
        resolved = manager.materialize_api_case(test_case)
        return resolved, manager
    except Exception as exc:
        raw = getattr(test_case, "test_data", None)
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except Exception:
                raw = {}
        raw = raw if isinstance(raw, dict) else {}
        requirements = raw.get("requirements", []) if raw else []
        has_mutation = any(isinstance(r, dict) and (r.get("mutation") or r.get("data_type") == "mutation")
                           for r in (requirements or []))
        if has_mutation:
            logger.error(f"API测试数据计划实例化失败，禁止回退原请求: case={getattr(test_case, 'id', '')}, error={exc}")
            return {"error": f"测试数据计划实例化失败: {exc}", "dataset": None, "plan": None}, None
        logger.warning(f"API测试数据计划实例化失败，用原请求继续: case={getattr(test_case, 'id', '')}, error={exc}")
        return {
            "query_params": dict(test_case.query_params or {}) if isinstance(test_case.query_params, dict) else {},
            "path_params": dict(test_case.path_params or {}) if isinstance(test_case.path_params, dict) else {},
            "headers": dict(test_case.headers or {}) if isinstance(test_case.headers, dict) else {},
            "request_body": test_case.request_body,
            "dataset": None, "plan": None,
        }, None


async def _execute_single_case_with_cache(
    test_case: ApiTestCase,
    base_url: str,
    db: Session,
    current_user: dict,
    credentials: Optional[HTTPAuthorizationCredentials],
    execution_cache: Dict[int, Dict[str, Any]],
    is_selected: bool,
    env_auth_vars: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """执行单个用例（带缓存）
    
    Args:
        test_case: 要执行的用例
        execution_cache: 已执行用例的缓存 {case_id: {extracted_vars, result}}
        is_selected: 是否是用户主动选择的用例
    
    Returns:
        {status, message, extracted_vars, skipped, actual_status}
    """
    from app.core.services.api_assert_executor import ApiAssertExecutor
    
    if test_case.id in execution_cache:
        logger.info(f"用例 {test_case.id} 已在缓存中，跳过执行")
        cached_result = execution_cache[test_case.id]
        return {
            "status": cached_result.get("status", "skipped"),
            "message": "已作为前置用例执行",
            "extracted_vars": cached_result.get("extracted_vars", {}),
            "skipped": True,
            "actual_status": cached_result.get("actual_status")
        }
    
    if not is_selected:
        depends_on = test_case.depends_on or []
        all_deps_executed = all(int(d) in execution_cache for d in depends_on if isinstance(depends_on, list))
        
        if not all_deps_executed:
            logger.info(f"前置用例 {test_case.id} 依赖未全部执行，开始执行")
    
    extracted_vars = {}
    if test_case.depends_on:
        depends_on = test_case.depends_on if isinstance(test_case.depends_on, list) else []
        for dep_id in depends_on:
            dep_id_int = int(dep_id)
            if dep_id_int in execution_cache:
                cached_vars = execution_cache[dep_id_int].get("extracted_vars", {})
                extracted_vars.update(cached_vars)
                logger.info(f"从缓存获取前置用例 {dep_id_int} 的变量: {list(cached_vars.keys())}")
            else:
                logger.warning(f"前置用例 {dep_id_int} 未在缓存中")
    
    runtime_request, data_manager = _resolve_api_runtime_request(test_case)
    if runtime_request and runtime_request.get("error"):
        return {
            "case_id": test_case.id, "name": test_case.name, "status": "error",
            "message": runtime_request["error"], "error_message": runtime_request["error"],
            "skipped": False, "extracted_vars": {},
        }
    runtime_dataset = runtime_request.get("dataset") if runtime_request else None
    runtime_plan = runtime_request.get("plan") if runtime_request else None
    _runtime_values = dict(runtime_request.get("dataset").values if runtime_request.get("dataset") else {})
    _runtime_values.update(extracted_vars)
    runtime_request["query_params"] = _render_runtime_structure(runtime_request.get("query_params") or {}, _runtime_values)
    runtime_request["path_params"] = _render_runtime_structure(runtime_request.get("path_params") or {}, _runtime_values)
    runtime_request["headers"] = _render_runtime_structure(runtime_request.get("headers") or {}, _runtime_values)
    runtime_request["request_body"] = _render_runtime_structure(runtime_request.get("request_body"), _runtime_values)

    if is_selected:
        execution = ApiTestExecution(
            case_id=test_case.id,
            project_id=test_case.project_id,
            status="running",
            start_time=datetime.utcnow()
        )
        db.add(execution)
        db.commit()
    
    path = test_case.path or "/"
    path_params_dict = runtime_request.get("path_params") or {}
    for key, value in path_params_dict.items():
        path = path.replace(f"{{{key}}}", str(value))
    if not path_params_dict:
        path = path.replace("{project_id}", "1")
        path = path.replace("{version_id}", "1")
        path = path.replace("{user_id}", "1")
        path = path.replace("{id}", "1")
        path = path.replace("{case_id}", str(test_case.id))
    
    url = f"{base_url}{path}"
    
    headers = dict(runtime_request.get("headers") or {})

    # 探索生成的 no_auth 变体用例：刻意不带鉴权（验证 401/403），跳过全部鉴权注入
    # ——定义在占位符替换之前，守卫语义完整覆盖下方全部注入分支（审计 H2）
    _skip_auth = isinstance(test_case.tags, list) and "no_auth" in test_case.tags

    # 探索生成用例的鉴权占位符替换：{{auth_token}} → 项目级 api_auth 实时 token
    # （探索生成时只落占位符不落明文 token；token 过期后重跑仍拿到当前有效值）
    if env_auth_vars and env_auth_vars.get("auth_token") and not _skip_auth:
        _tok = env_auth_vars["auth_token"]
        _inj = env_auth_vars.get("token_injection") or {}
        if isinstance(_inj, str):
            try:
                _inj = json.loads(_inj)
            except Exception:
                _inj = {}
        _prefix = (_inj.get("prefix") if isinstance(_inj, dict) else None) or "Bearer "
        for _hk in list(headers.keys()):
            _hv = headers[_hk]
            if isinstance(_hv, str) and "{{auth_token}}" in _hv:
                headers[_hk] = _hv.replace("{{auth_token}}", f"{_prefix}{_tok}")
                logger.info(f"鉴权占位符已替换为运行时token: {_hk}")
    # 无 token 可用时：移除残留占位符头（避免把字面 {{auth_token}} 发出去）
    for _hk in list(headers.keys()):
        if isinstance(headers[_hk], str) and "{{auth_token}}" in headers[_hk]:
            headers.pop(_hk, None)
            logger.warning(f"鉴权占位符无 token 可替换，已移除请求头: {_hk}")

    # 环境鉴权变量合并
    if env_auth_vars and not _skip_auth:
        # API Key注入
        if env_auth_vars.get("api_keys"):
            headers.update(env_auth_vars["api_keys"])
        # Cookie注入
        if env_auth_vars.get("cookies"):
            cookie_str = "; ".join(f"{k}={v}" for k, v in env_auth_vars["cookies"].items())
            headers["Cookie"] = cookie_str
        # auth_header注入 (Basic Auth)
        if env_auth_vars.get("auth_header"):
            headers["Authorization"] = env_auth_vars["auth_header"]

    # Token优先级: 1. 前置用例提取 2. 环境鉴权 3. 用户凭证（no_auth 变体跳过）
    if not _skip_auth and extracted_vars.get("auth_token"):
        headers["Authorization"] = f"Bearer {extracted_vars['auth_token']}"
        logger.info(f"使用缓存的token执行用例 {test_case.id}")
    elif not _skip_auth and env_auth_vars and env_auth_vars.get("auth_token"):
        token = env_auth_vars["auth_token"]
        inj = (env_auth_vars.get("token_injection") or {})
        if isinstance(inj, str):
            try:
                inj = json.loads(inj)
            except Exception:
                inj = {}
        if isinstance(inj, dict):
            header_name = inj.get("header_name", "Authorization")
            prefix = inj.get("prefix", "Bearer ")
            headers[header_name] = f"{prefix}{token}"
        else:
            headers["Authorization"] = f"Bearer {token}"
        logger.info(f"使用环境鉴权token执行用例 {test_case.id}")
    elif not _skip_auth and test_case.case_type != "auth" and credentials:
        user_token = credentials.credentials
        headers["Authorization"] = f"Bearer {user_token}"
    
    params = dict(runtime_request.get("query_params") or {})
    
    body = runtime_request.get("request_body")
    if isinstance(body, dict) and not body:
        body = None
    elif body is None and test_case.request_body:
        if isinstance(test_case.request_body, dict):
            body = dict(test_case.request_body)
        elif isinstance(test_case.request_body, str):
            # A3 补全（2026-08-25）：非 JSON 原文/畸形 JSON 字符串 body 此前在提取端
            # 被直接丢弃 → 发送端 str→data= 分支永不可达（修复只做了发送端一半）。
            # 原样保留，发送时按 data= 传递（见 is_json_content 分支）。
            body = test_case.request_body

    # 特殊处理：注册接口每次执行动态生成随机用户名，避免重复注册失败
    is_register_endpoint = "register" in test_case.path.lower() or "signup" in test_case.path.lower()
    if is_register_endpoint and isinstance(body, dict) and body and test_case.method == "POST":
        import random
        import time
        timestamp = int(time.time() * 1000) % 100000
        random_suffix = random.randint(1000, 9999)
        
        if "username" in body:
            body["username"] = f"testuser_{timestamp}_{random_suffix}"
            logger.info(f"注册接口动态生成用户名: {body['username']}")
        
        if "email" in body:
            body["email"] = f"test_{timestamp}_{random_suffix}@example.com"
            logger.info(f"注册接口动态生成邮箱: {body['email']}")
        
        if "password" in body and "confirm_password" in body:
            body["confirm_password"] = body["password"]
            logger.info(f"注册接口同步confirm_password")
    
    # 变量替换：将请求体中的示例值替换为实际提取的变量值
    # 如果请求体中有字段值是示例值（如 test_value, test_token_xxx），且对应变量存在，则替换
    if isinstance(body, dict) and body and extracted_vars:
        for key, value in body.items():
            # 如果值是示例值格式的字符串，尝试从 extracted_vars 获取同名变量
            if isinstance(value, str) and value.startswith('test_'):
                # 尝试匹配变量名（去除 test_ 前缀）
                var_name = key  # 直接使用字段名作为变量名
                if var_name in extracted_vars:
                    body[key] = extracted_vars[var_name]
                    logger.info(f"变量替换: {key} 从 '{value}' 替换为提取的值")
                # 也尝试常见的变体名（如 refresh_token -> refresh_token）
                if key in ['token', 'access_token', 'auth_token'] and 'auth_token' in extracted_vars:
                    body[key] = extracted_vars['auth_token']
                    logger.info(f"变量替换: {key} 替换为 auth_token")
    
    logger.info(f"执行用例 {test_case.id} ({test_case.name}): {test_case.method} {url}")
    
    start_time = datetime.utcnow()
    
    result = {
        "case_id": test_case.id,
        "name": test_case.name,
        "extracted_vars": {},
        "skipped": False,
        "method": str(test_case.method or "GET"),
        "request_url": url,
        "request_headers": headers,
        "request_params": params,
        "request_body": body
    }
    
    # 确定请求体的发送方式：根据接口类型选择 json 或 data
    # OAuth2 表单格式接口（如 /auth/login）需要用 data 参数
    # JSON 格式接口（如 /auth/login/json）需要用 json 参数
    
    # 检查接口路径判断是否是OAuth2表单格式
    oauth2_form_paths = ['/auth/login', '/auth/token', '/login', '/token', '/signin']
    is_oauth2_form = any(p in test_case.path for p in oauth2_form_paths) and test_case.method == 'POST'
    
    # 如果headers中有Content-Type，优先使用它来判断
    content_type = headers.get('Content-Type', headers.get('content-type', ''))
    is_json_content = 'application/json' in content_type.lower() if content_type else not is_oauth2_form
    
    logger.debug(f"用例 {test_case.id}: path={test_case.path}, content_type={content_type}, is_oauth2_form={is_oauth2_form}, is_json_content={is_json_content}")
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            # 检查请求体类型：multipart/form-data
            body_type = body.get("_body_type", "json") if isinstance(body, dict) else "json"

            if body_type == "multipart":
                # 构建 multipart/form-data 请求
                files = {}
                data_fields = {}
                multipart_fields = body.get("fields", []) if isinstance(body, dict) else []

                for field in multipart_fields if isinstance(multipart_fields, list) else []:
                    field_name = field.get("name", "")
                    field_type = field.get("type", "text")

                    if field_type == "file":
                        file_path = field.get("file_path", "")
                        mime_type = field.get("mime_type", "application/octet-stream")
                        file_name = field.get("file_name", "")

                        if file_path and os.path.exists(file_path):
                            files[field_name] = (file_name or os.path.basename(file_path),
                                                 open(file_path, "rb"), mime_type)
                        elif field.get("content"):
                            import base64 as _base64
                            file_content = _base64.b64decode(field["content"])
                            files[field_name] = (file_name or "file", file_content, mime_type)
                    else:
                        data_fields[field_name] = field.get("value", "")

                response = await client.request(
                    method=str(test_case.method or "GET"),
                    url=url,
                    headers=headers if headers else None,
                    params=params if params else None,
                    data=data_fields if data_fields else None,
                    files=files if files else None,
                )

                # 关闭文件句柄
                for k, v in files.items():
                    if isinstance(v, tuple) and len(v) >= 2 and hasattr(v[1], 'close'):
                        try:
                            v[1].close()
                        except Exception:
                            pass
            elif is_json_content:
                # JSON 格式请求：dict 走 json=（httpx 序列化）；字符串 body（探索捕获的
                # 非 JSON 原文/畸形 JSON 保留形态）按 data= 原样发送——json= 传字符串
                # 会被 JSON 编码改变语义（A3 修复 2026-08-25）
                response = await client.request(
                    method=str(test_case.method or "GET"),
                    url=url,
                    headers=headers if headers else None,
                    params=params if params else None,
                    json=body if isinstance(body, dict) and body else None,
                    data=body if isinstance(body, str) else None,
                )
            else:
                # 表单格式请求（OAuth2）
                # 如果headers中没有Content-Type，添加表单类型
                if not content_type:
                    headers['Content-Type'] = 'application/x-www-form-urlencoded'
                response = await client.request(
                    method=str(test_case.method or "GET"),
                    url=url,
                    headers=headers if headers else None,
                    params=params if params else None,
                    data=body if body else None
                )
            
            actual_status = response.status_code
            result["actual_status"] = actual_status
            result["response_headers"] = dict(response.headers) if response.headers else {}
            result["duration"] = max(0, int((datetime.utcnow() - start_time).total_seconds() * 1000))
            
            try:
                response_body = response.json()
                result["response_body"] = response_body
            except:
                response_body = {}
                result["response_body"] = response.text[:500] if response.text else {}
            
            if test_case.variable_extractions:
                for extraction in test_case.variable_extractions:
                    var_name = extraction.get("name")
                    source = extraction.get("source", "response_body")
                    path_ext = extraction.get("path", "")
                    
                    if source == "response_body" and path_ext:
                        value = _extract_value_from_dict(response_body, path_ext)
                        if value:
                            extracted_vars[var_name] = value
                            logger.info(f"提取变量: {var_name} = {str(value)[:20]}...")
            
            if not extracted_vars.get("auth_token"):
                token_fields = ["token", "access_token", "auth_token", "data.token", "data.access_token"]
                for field in token_fields:
                    value = _extract_value_from_dict(response_body, field)
                    if value:
                        extracted_vars["auth_token"] = value
                        logger.info(f"自动提取token: {field}")
                        break
            
            result["extracted_vars"] = extracted_vars
            
            assert_rules = []
            if test_case.assert_rules:
                if isinstance(test_case.assert_rules, list):
                    assert_rules = test_case.assert_rules
                elif isinstance(test_case.assert_rules, dict):
                    assert_rules = [test_case.assert_rules]
            
            expected_http_status_codes = None
            for rule in assert_rules:
                if rule.get("type") == "http_status" and rule.get("value"):
                    expected_http_status_codes = rule.get("value")
                    break
            
            if expected_http_status_codes is None:
                expected_http_status_codes = [test_case.expected_status or 200]
            
            http_passed = actual_status in expected_http_status_codes
            
            if assert_rules and response_body:
                assert_executor = ApiAssertExecutor(response_body, assert_rules)
                assert_executor.execute()
                all_passed = http_passed and assert_executor.is_all_passed()
                result["assert_results"] = assert_executor.results
            else:
                all_passed = http_passed
                result["assert_results"] = []
            
            if all_passed:
                result["status"] = "passed"
                result["message"] = "执行成功"
            else:
                result["status"] = "failed"
                error_msgs = []
                if not http_passed:
                    error_msgs.append(f"HTTP状态码不匹配: 期望 {expected_http_status_codes}, 实际 {actual_status}")
                if result.get("assert_results"):
                    for ar in result["assert_results"]:
                        if not ar.get("passed"):
                            error_msgs.append(ar.get("message", "断言失败"))
                result["message"] = "; ".join(error_msgs) if error_msgs else "断言失败"
                result["error_message"] = result["message"]
            
            if is_selected:
                execution.actual_status = actual_status
                execution.actual_body = response_body
                execution.actual_headers = dict(response.headers) if response.headers else {}
                execution.status = result["status"]
                execution.end_time = datetime.utcnow()
                execution.duration = max(0, int((execution.end_time - execution.start_time).total_seconds() * 1000))
                if result["status"] == "failed":
                    execution.error_message = result["message"]
                db.commit()
    
    except Exception as e:
        result["status"] = "error"
        result["message"] = str(e)
        logger.error(f"执行用例 {test_case.id} 失败: {str(e)}")
        
        if is_selected:
            execution.status = "error"
            execution.error_message = str(e)
            execution.end_time = datetime.utcnow()
            execution.duration = max(0, int((execution.end_time - execution.start_time).total_seconds() * 1000))
            db.commit()
    
    if data_manager and runtime_dataset and runtime_plan:
        try:
            if result.get("status") in ("passed", "failed"):
                for req in runtime_plan.requirements:
                    if req.data_type == "consumable":
                        data_manager.lifecycle.mark_consumed(runtime_dataset, req.key, {"case_id": test_case.id, "status": result.get("status")})
            result["data_set_id"] = runtime_dataset.run_id
            result["test_data_plan"] = runtime_plan.to_dict()
            result["data_cleanup"] = data_manager.lifecycle.complete(runtime_dataset, runtime_plan, data_manager)
        except Exception as _data_e:
            logger.warning(f"API测试数据生命周期处理失败 case={test_case.id}: {_data_e}")
    execution_cache[test_case.id] = result
    
    return result


@router.post("/execute", response_model=ApiTestExecutionResponse)
async def execute_test(
    request: ExecuteApiTestRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False)),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """执行API测试，验证业务响应码和数据结构"""
    from app.core.services.api_assert_executor import ApiAssertExecutor
    
    test_case = db.query(ApiTestCase).filter(ApiTestCase.id == request.case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"用例ID {request.case_id} 不存在")
    
    execution = ApiTestExecution(
        case_id=test_case.id,
        project_id=test_case.project_id,
        environment=request.environment,
        trigger_type="manual",
        trigger_user_id=str(current_user["user"].id),
        status="running",
        start_time=datetime.utcnow()
    )
    
    db.add(execution)
    db.commit()
    
    # 用例 base_url 优先（探索生成=捕获的真实被测 origin；Swagger 导入=接口文档地址）；
    # request.base_url（前端代理地址）仅作无 base_url 老用例的兜底（审计 M2）
    base_url = normalize_url_colon(test_case.base_url or request.base_url or "http://localhost:8000")
    
    if base_url and 'localhost' in base_url and ':' not in base_url.split('/')[-1]:
        if 'http://localhost' in base_url:
            base_url = 'http://localhost:8000'
        elif 'https://localhost' in base_url:
            base_url = 'https://localhost:8443'
    
    # 环境鉴权
    env_auth_vars = None
    try:
        env = db.query(ApiEnvironment).filter(
            ApiEnvironment.project_id == test_case.project_id,
            ApiEnvironment.is_default == True
        ).first()
        # env 可能为 None（项目未建默认环境）→ 显式传 project_id 使项目级 api_auth 回退仍可用
        env_auth_vars = await _execute_env_auth(env, base_url, db=db, project_id=test_case.project_id)
        if env_auth_vars and env_auth_vars.get("logs"):
            for log_msg in env_auth_vars["logs"]:
                logger.info(log_msg)
    except Exception as e:
        logger.warning(f"环境鉴权失败: {str(e)}")

    # 执行前置用例，获取变量（如auth_token）
    extracted_vars = {}
    if test_case.depends_on:
        try:
            extracted_vars = await _execute_precondition_cases(
                test_case.depends_on,
                base_url,
                db,
                current_user,
                credentials,
                env_auth_vars=env_auth_vars
            )
            logger.info(f"Extracted variables from precondition cases: {extracted_vars.keys()}")
        except Exception as e:
            logger.warning(f"Failed to execute precondition cases: {str(e)}")

    runtime_request, data_manager = _resolve_api_runtime_request(test_case)
    if runtime_request and runtime_request.get("error"):
        return {
            "case_id": test_case.id, "name": test_case.name, "status": "error",
            "message": runtime_request["error"], "error_message": runtime_request["error"],
            "skipped": False, "extracted_vars": {},
        }
    runtime_dataset = runtime_request.get("dataset") if runtime_request else None
    runtime_plan = runtime_request.get("plan") if runtime_request else None
    _runtime_values = dict(runtime_request.get("dataset").values if runtime_request.get("dataset") else {})
    _runtime_values.update(extracted_vars)
    runtime_request["query_params"] = _render_runtime_structure(runtime_request.get("query_params") or {}, _runtime_values)
    runtime_request["path_params"] = _render_runtime_structure(runtime_request.get("path_params") or {}, _runtime_values)
    runtime_request["headers"] = _render_runtime_structure(runtime_request.get("headers") or {}, _runtime_values)
    runtime_request["request_body"] = _render_runtime_structure(runtime_request.get("request_body"), _runtime_values)

    path = test_case.path or "/"
    
    if runtime_request.get("path_params"):
        for key, value in (runtime_request.get("path_params") or {}).items():
            path = path.replace(f"{{{key}}}", str(value))
    else:
        path = path.replace("{project_id}", "1")
        path = path.replace("{version_id}", "1")
        path = path.replace("{user_id}", "1")
        path = path.replace("{id}", "1")
        path = path.replace("{case_id}", str(test_case.id))
    
    url = f"{base_url}{path}"
    
    headers = dict(runtime_request.get("headers") or {})

    # 探索生成的 no_auth 变体用例：刻意不带鉴权（验证 401/403），跳过全部鉴权注入
    # ——定义在占位符替换之前，守卫语义完整覆盖下方全部注入分支（审计 H2）
    _skip_auth = isinstance(test_case.tags, list) and "no_auth" in test_case.tags

    # 探索生成用例的鉴权占位符替换：{{auth_token}} → 项目级 api_auth 实时 token
    # （探索生成时只落占位符不落明文 token；token 过期后重跑仍拿到当前有效值）
    if env_auth_vars and env_auth_vars.get("auth_token") and not _skip_auth:
        _tok = env_auth_vars["auth_token"]
        _inj = env_auth_vars.get("token_injection") or {}
        if isinstance(_inj, str):
            try:
                _inj = json.loads(_inj)
            except Exception:
                _inj = {}
        _prefix = (_inj.get("prefix") if isinstance(_inj, dict) else None) or "Bearer "
        for _hk in list(headers.keys()):
            _hv = headers[_hk]
            if isinstance(_hv, str) and "{{auth_token}}" in _hv:
                headers[_hk] = _hv.replace("{{auth_token}}", f"{_prefix}{_tok}")
                logger.info(f"鉴权占位符已替换为运行时token: {_hk}")
    # 无 token 可用时：移除残留占位符头（避免把字面 {{auth_token}} 发出去）
    for _hk in list(headers.keys()):
        if isinstance(headers[_hk], str) and "{{auth_token}}" in headers[_hk]:
            headers.pop(_hk, None)
            logger.warning(f"鉴权占位符无 token 可替换，已移除请求头: {_hk}")

    # 环境鉴权变量合并
    if env_auth_vars and not _skip_auth:
        if env_auth_vars.get("api_keys"):
            headers.update(env_auth_vars["api_keys"])
        if env_auth_vars.get("cookies"):
            cookie_str = "; ".join(f"{k}={v}" for k, v in env_auth_vars["cookies"].items())
            headers["Cookie"] = cookie_str
        if env_auth_vars.get("auth_header"):
            headers["Authorization"] = env_auth_vars["auth_header"]

    # 使用从前置用例提取的auth_token（no_auth 变体跳过）
    if not _skip_auth and extracted_vars.get("auth_token"):
        headers["Authorization"] = f"Bearer {extracted_vars['auth_token']}"
    elif not _skip_auth and env_auth_vars and env_auth_vars.get("auth_token"):
        token = env_auth_vars["auth_token"]
        inj = (env_auth_vars.get("token_injection") or {})
        if isinstance(inj, str):
            try:
                inj = json.loads(inj)
            except Exception:
                inj = {}
        if isinstance(inj, dict):
            header_name = inj.get("header_name", "Authorization")
            prefix = inj.get("prefix", "Bearer ")
            headers[header_name] = f"{prefix}{token}"
        else:
            headers["Authorization"] = f"Bearer {token}"
    elif not _skip_auth and test_case.case_type != "auth" and credentials:
        # 如果没有前置提取的token，但不是auth类型测试，使用当前用户token
        user_token = credentials.credentials
        headers["Authorization"] = f"Bearer {user_token}"
    
    params = dict(runtime_request.get("query_params") or {})
    
    body = runtime_request.get("request_body")
    request_body_raw = body if body is not None else test_case.request_body
    logger.info(f"request_body from database: {request_body_raw}, type: {type(request_body_raw)}")

    if request_body_raw:
        if isinstance(request_body_raw, dict):
            body = dict(request_body_raw)
            logger.info(f"body will be sent: {body}")
        elif isinstance(request_body_raw, str):
            # 非 JSON 原文 body（探索捕获的 form/text，A3 修复 2026-08-25——此前字符串
            # body 被直接丢弃，用例请求缺失 body 必失败）：截断保留，下方按 data= 发送
            body = request_body_raw[:_RAW_BODY_MAX_CHARS]
    
    logger.info(f"Executing API test: {test_case.method} {url}, case_type: {test_case.case_type}, body={body}")
    
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.request(
                method=str(test_case.method or "GET"),
                url=url,
                headers=headers if headers else None,
                params=params if params else None,
                # dict → json=；字符串原文（非 JSON body，A3 修复）→ data= 原样发送
                json=body if isinstance(body, dict) and body else None,
                data=body if isinstance(body, str) else None,
            )
            
            execution.actual_status = response.status_code
            
            try:
                response_body = response.json()
                execution.actual_body = response_body
            except:
                response_body = {}
                execution.actual_body = {"raw": response.text[:1000] if response.text else ""}
            
            execution.actual_headers = dict(response.headers) if response.headers else {}
            
            assert_rules = []
            if test_case.assert_rules:
                if isinstance(test_case.assert_rules, list):
                    assert_rules = test_case.assert_rules
                elif isinstance(test_case.assert_rules, dict):
                    assert_rules = [test_case.assert_rules]
            
            # 从断言规则中提取期望的HTTP状态码列表
            expected_http_status_codes = None
            for rule in assert_rules:
                if rule.get("type") == "http_status" and rule.get("value"):
                    expected_http_status_codes = rule.get("value")
                    break
            
            # 如果断言规则中没有指定HTTP状态码，使用用例的expected_status
            if expected_http_status_codes is None:
                expected_http_status_codes = [test_case.expected_status or 200]
            
            http_passed = response.status_code in expected_http_status_codes
            
            if assert_rules and response_body:
                assert_executor = ApiAssertExecutor(response_body, assert_rules)
                assert_results = assert_executor.execute()
                execution.assert_results = [
                    {
                        "rule": r.get("rule_type", ""),
                        "passed": r.get("passed", False),
                        "message": r.get("message", "")
                    }
                    for r in assert_results
                ]
                
                if http_passed and assert_executor.is_all_passed():
                    execution.status = "passed"
                else:
                    execution.status = "failed"
                    
            else:
                execution.status = "passed" if http_passed else "failed"
            
            if execution.status == "failed":
                error_messages = []
                if not http_passed:
                    error_messages.append(f"HTTP状态码不匹配: 期望 {expected_http_status_codes}, 实际 {response.status_code}")
                
                if response_body and "code" in response_body:
                    error_messages.append(f"业务返回码: {response_body.get('code')}")
                
                if execution.assert_results:
                    failed_asserts = [r for r in execution.assert_results if not r.get("passed")]
                    if failed_asserts:
                        error_messages.extend([r.get("message") for r in failed_asserts])
                
                execution.error_message = "; ".join(error_messages)
            
    except Exception as e:
        execution.status = "error"
        execution.error_message = str(e)
    
    execution.end_time = datetime.utcnow()
    duration_ms = int((execution.end_time - execution.start_time).total_seconds() * 1000)
    execution.duration = max(0, duration_ms)  # 确保duration不为负数
    
    if data_manager and runtime_dataset and runtime_plan:
        try:
            if execution.status in ("passed", "failed"):
                for req in runtime_plan.requirements:
                    if req.data_type == "consumable":
                        data_manager.lifecycle.mark_consumed(runtime_dataset, req.key, {"case_id": test_case.id, "status": execution.status})
            db.commit()
            data_manager.lifecycle.complete(runtime_dataset, runtime_plan, data_manager)
        except Exception as _data_e:
            logger.warning(f"API测试数据生命周期处理失败 case={test_case.id}: {_data_e}")

    db.commit()
    db.refresh(execution)
    
    logger.info(f"执行API测试: {test_case.name}, 状态: {execution.status}")
    
    return ApiTestExecutionResponse.model_validate(execution)


@router.get("/environments/{project_id}", response_model=List[ApiEnvironmentResponse])
def list_environments(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取项目的环境列表"""
    environments = db.query(ApiEnvironment).filter(
        ApiEnvironment.project_id == project_id
    ).order_by(ApiEnvironment.is_default.desc()).all()
    
    return [ApiEnvironmentResponse.model_validate(e) for e in environments]


@router.post("/environments", response_model=ApiEnvironmentResponse, status_code=status.HTTP_201_CREATED)
def create_environment(
    env_in: ApiEnvironmentCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """创建测试环境"""
    if env_in.is_default:
        db.query(ApiEnvironment).filter(
            ApiEnvironment.project_id == env_in.project_id,
            ApiEnvironment.is_default == True
        ).update({"is_default": False})
    
    environment = ApiEnvironment(**env_in.model_dump())
    
    db.add(environment)
    db.commit()
    db.refresh(environment)
    
    return ApiEnvironmentResponse.model_validate(environment)


@router.post("/auto-generate", response_model=SwaggerAutoGenerateResponse, status_code=status.HTTP_201_CREATED)
async def auto_generate_from_swagger(
    request: SwaggerAutoGenerateRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """
    从Swagger URL自动生成API测试用例
    """
    from app.core.services.api_test_generator import ApiTestGeneratorService
    import traceback
    
    try:
        logger.info(f"Auto-generate request: project_id={request.project_id}, version_id={request.version_id}, swagger_url={request.swagger_url}")
        
        generator = ApiTestGeneratorService(db)
        
        result = await generator.auto_generate_from_swagger(
            project_id=request.project_id,
            version_id=request.version_id,
            swagger_url=request.swagger_url,
            base_url=request.base_url,
            include_normal=request.include_normal,
            include_error=request.include_error,
            include_boundary=request.include_boundary,
            include_auth=request.include_auth,
            max_cases_per_endpoint=request.max_cases_per_endpoint,
            user_id=current_user["user"].id
        )
        
        logger.info(f"Generation result: success={result.get('success')}, endpoints={result.get('endpoints_count')}, cases={result.get('generated_count')}")
        
        test_cases_response = []
        for tc in result.get("test_cases", []):
            try:
                test_cases_response.append(GeneratedApiTestCase(
                    id=tc.id,
                    name=tc.name,
                    endpoint_path=tc.path or "",
                    method=tc.method or "GET",
                    case_type=tc.case_type or "normal",
                    priority=tc.priority or "P2",
                    description=tc.description,
                    preconditions=tc.preconditions if hasattr(tc, 'preconditions') else None,
                    test_steps=tc.test_steps if hasattr(tc, 'test_steps') else None,
                    expected_result=tc.expected_result if hasattr(tc, 'expected_result') else None,
                    headers=tc.headers,
                    query_params=tc.query_params,
                    request_body=tc.request_body,
                    test_data=getattr(tc, "test_data", None),
                    expected_status=tc.expected_status,
                    assert_rules=tc.assert_rules
                ))
            except Exception as e:
                logger.warning(f"Failed to create response for test case: {str(e)}")
                logger.warning(traceback.format_exc())
        
        return SwaggerAutoGenerateResponse(
            success=result["success"],
            message=result["message"],
            definition_id=result.get("definition_id"),
            endpoints_count=result["endpoints_count"],
            generated_count=result["generated_count"],
            test_cases=test_cases_response,
            generation_summary=result.get("generation_summary"),
            raw_spec=result.get("raw_spec"),
        )
    except KeyError as e:
        logger.error(f"KeyError in auto-generate: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"KeyError: {str(e)}")
    except Exception as e:
        logger.error(f"Auto-generate failed: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/cases/version/{version_id}", response_model=ApiTestCaseListResponse)
def list_test_cases_by_version(
    version_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    case_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取版本下的API测试用例列表"""
    query = db.query(ApiTestCase).filter(ApiTestCase.version_id == version_id)
    
    if case_type:
        query = query.filter(ApiTestCase.case_type == case_type)
    if priority:
        query = query.filter(ApiTestCase.priority == priority)
    if search:
        pattern = f"%{search}%"
        query = query.filter(ApiTestCase.name.ilike(pattern))
    
    total = query.count()
    total_pages = (total + page_size - 1) // page_size
    
    cases = query.order_by(ApiTestCase.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return ApiTestCaseListResponse(
        items=[ApiTestCaseResponse.model_validate(c) for c in cases],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.get("/cases/project/{project_id}", response_model=ApiTestCaseListResponse)
def list_test_cases_by_project(
    project_id: int,
    include_unclassified: bool = Query(True, description="是否包含未分类用例(version_id为空)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    case_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取项目下的所有API测试用例（包含未分类的）"""
    query = db.query(ApiTestCase).filter(ApiTestCase.project_id == project_id)
    
    if include_unclassified:
        # 包含所有用例，不管version_id
        pass
    else:
        # 只包含有version_id的用例
        query = query.filter(ApiTestCase.version_id.isnot(None))
    
    if case_type:
        query = query.filter(ApiTestCase.case_type == case_type)
    if priority:
        query = query.filter(ApiTestCase.priority == priority)
    if search:
        pattern = f"%{search}%"
        query = query.filter(ApiTestCase.name.ilike(pattern))
    
    total = query.count()
    total_pages = (total + page_size - 1) // page_size
    
    cases = query.order_by(ApiTestCase.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return ApiTestCaseListResponse(
        items=[ApiTestCaseResponse.model_validate(c) for c in cases],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.get("/cases/unclassified/{project_id}", response_model=ApiTestCaseListResponse)
def list_unclassified_test_cases(
    project_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    case_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取项目下未分类的API测试用例（version_id为空）"""
    query = db.query(ApiTestCase).filter(
        ApiTestCase.project_id == project_id,
        ApiTestCase.version_id.is_(None)
    )
    
    if case_type:
        query = query.filter(ApiTestCase.case_type == case_type)
    if priority:
        query = query.filter(ApiTestCase.priority == priority)
    if search:
        pattern = f"%{search}%"
        query = query.filter(ApiTestCase.name.ilike(pattern))
    
    total = query.count()
    total_pages = (total + page_size - 1) // page_size
    
    cases = query.order_by(ApiTestCase.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return ApiTestCaseListResponse(
        items=[ApiTestCaseResponse.model_validate(c) for c in cases],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.get("/cases/{case_id}", response_model=ApiTestCaseResponse)
def get_test_case_detail(
    case_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取API测试用例详情"""
    test_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"用例ID {case_id} 不存在")
    
    return ApiTestCaseResponse.model_validate(test_case)


@router.put("/cases/{case_id}", response_model=ApiTestCaseResponse)
def update_test_case(
    case_id: int,
    case_in: ApiTestCaseUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """更新API测试用例"""
    test_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"用例ID {case_id} 不存在")
    
    update_data = case_in.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(test_case, key, value)
    
    db.commit()
    db.refresh(test_case)
    
    logger.info(f"更新API测试用例: {test_case.name}")
    
    return ApiTestCaseResponse.model_validate(test_case)


@router.delete("/cases/{case_id}")
def delete_test_case(
    case_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """删除API测试用例（同时删除相关执行记录）"""
    test_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"用例ID {case_id} 不存在")
    
    # 先删除相关的执行记录
    db.query(ApiTestExecution).filter(ApiTestExecution.case_id == case_id).delete()
    
    db.delete(test_case)
    db.commit()
    
    logger.info(f"删除API测试用例: {case_id}，及相关执行记录")
    
    return {"message": "删除成功"}


@router.get("/executions")
def list_executions(
    case_id: Optional[int] = Query(None, description="用例ID"),
    project_id: Optional[int] = Query(None, description="项目ID"),
    status: Optional[str] = Query(None, description="执行状态"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取测试执行历史"""
    query = db.query(ApiTestExecution)
    
    if case_id:
        query = query.filter(ApiTestExecution.case_id == case_id)
    if project_id:
        query = query.filter(ApiTestExecution.project_id == project_id)
    if status:
        query = query.filter(ApiTestExecution.status == status)
    
    total = query.count()
    executions = query.order_by(ApiTestExecution.start_time.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return {
        "items": [ApiTestExecutionResponse.model_validate(e) for e in executions],
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.post("/cases/batch-delete")
def batch_delete_test_cases(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """批量删除API测试用例（同时删除相关执行记录）"""
    if not request.case_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的用例")
    
    deleted_count = 0
    for case_id in request.case_ids:
        # 先删除相关的执行记录
        db.query(ApiTestExecution).filter(ApiTestExecution.case_id == case_id).delete()
        
        # 再删除用例
        test_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id).first()
        if test_case:
            db.delete(test_case)
            deleted_count += 1
    
    db.commit()
    
    logger.info(f"批量删除API测试用例: {deleted_count} 条，及相关执行记录")
    
    return {"message": f"成功删除 {deleted_count} 条用例及相关执行记录", "deleted_count": deleted_count}


@router.post("/cases/batch-execute", response_model=BatchExecuteResponse)
async def batch_execute_test_cases(
    request: BatchExecuteRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False)),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """批量执行API测试用例（支持依赖处理和执行缓存）
    
    执行流程：
    1. 获取所有用例信息
    2. 拓扑排序（根据depends_on关系）
    3. 按顺序执行，前置用例优先
    4. 缓存已执行用例的结果（如token）
    5. 后续用例复用缓存的变量
    
    特点：
    - 登录用例作为前置用例执行后，缓存token供后续用例使用
    - 用户勾选登录+其他用例时，登录先执行，其他用例复用token
    - 用户只勾选其他用例时，自动执行依赖的前置用例（登录）
    - 前置用例执行后自动跳过，不重复执行
    """
    if not request.case_ids:
        raise HTTPException(status_code=400, detail="请选择要执行的用例")
    
    selected_cases = []
    missing_ids = []
    
    for case_id in request.case_ids:
        test_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id).first()
        if test_case:
            selected_cases.append(test_case)
        else:
            missing_ids.append(case_id)
    
    if missing_ids:
        logger.warning(f"部分用例不存在: {missing_ids}")
    
    if not selected_cases:
        raise HTTPException(status_code=400, detail="没有有效的用例可执行")
    
    project_ids = {case.project_id for case in selected_cases}
    if len(project_ids) > 1:
        logger.warning(f"用例跨多个项目: {project_ids}")
    
    all_cases = db.query(ApiTestCase).filter(ApiTestCase.project_id.in_(project_ids)).all()
    all_cases_dict = {case.id: case for case in all_cases}
    
    sorted_cases, selected_ids = _topological_sort_cases(selected_cases, all_cases_dict)
    
    logger.info(f"批量执行: 用户选择 {len(selected_cases)} 个用例，拓扑排序后共 {len(sorted_cases)} 个（含前置）")
    logger.info(f"用户选择用例ID: {list(selected_ids)}")
    logger.info(f"排序后执行顺序: {[c.id for c in sorted_cases]}")
    
    base_url = request.base_url
    env = None
    if not base_url:
        env = db.query(ApiEnvironment).filter(
            ApiEnvironment.project_id.in_(project_ids),
            ApiEnvironment.is_default == True
        ).first()
        base_url = env.base_url if env else "http://localhost:8000"
    else:
        # 如果指定了base_url，也尝试查找环境用于鉴权
        env = db.query(ApiEnvironment).filter(
            ApiEnvironment.project_id.in_(project_ids),
            ApiEnvironment.is_default == True
        ).first()

    # 环境鉴权（env 为 None 时仍会回退项目级 api_auth——需显式传 project_id，与单条路径一致）
    env_auth_vars = None
    env_auth_vars = await _execute_env_auth(
        env, base_url, db=db,
        project_id=sorted(project_ids)[0] if project_ids else None,
    )
    if env_auth_vars and env_auth_vars.get("logs"):
        for log_msg in env_auth_vars["logs"]:
            logger.info(log_msg)
    
    if base_url and 'localhost' in base_url and ':' not in base_url.split('/')[-1]:
        if 'http://localhost' in base_url:
            base_url = 'http://localhost:8000'
        elif 'https://localhost' in base_url:
            base_url = 'https://localhost:8443'
    
    logger.info(f"使用base_url: {base_url}")
    
    execution_cache: Dict[int, Dict[str, Any]] = {}
    
    results = []
    passed_count = 0
    failed_count = 0
    error_count = 0
    skipped_count = 0
    precondition_count = 0
    
    for test_case in sorted_cases:
        is_selected = test_case.id in selected_ids
        
        if not is_selected:
            precondition_count += 1
            logger.info(f"执行前置用例 {test_case.id} ({test_case.name})")
        
        result = await _execute_single_case_with_cache(
            test_case=test_case,
            # 每条用例用自己的 base_url（真实被测地址），无则回退批量兜底值（审计 M2）
            base_url=test_case.base_url or base_url,
            db=db,
            current_user=current_user,
            credentials=credentials,
            execution_cache=execution_cache,
            is_selected=is_selected,
            env_auth_vars=env_auth_vars,
        )
        
        if result.get("skipped"):
            skipped_count += 1
            if is_selected:
                passed_count += 1
                results.append({
                    "case_id": test_case.id,
                    "name": test_case.name,
                    "status": "passed",
                    "message": "已作为前置用例执行",
                    "skipped": True,
                    "actual_status": result.get("actual_status")
                })
            continue
        
        if is_selected:
            status = result.get("status", "error")
            if status == "passed":
                passed_count += 1
            elif status == "failed":
                failed_count += 1
            else:
                error_count += 1
            
            results.append({
                "case_id": test_case.id,
                "name": test_case.name,
                "status": status,
                "message": result.get("message", ""),
                "actual_status": result.get("actual_status"),
                "duration": result.get("duration"),
                "skipped": False,
                "method": test_case.method,
                "request_url": result.get("request_url"),
                "request_headers": result.get("request_headers"),
                "request_params": result.get("request_params") or test_case.query_params,
                "request_body": result.get("request_body") or test_case.request_body,
                "response_headers": result.get("response_headers"),
                "response_body": result.get("response_body"),
                "error_message": result.get("error_message"),
                "assert_results": result.get("assert_results")
            })
    
    total_selected = len(selected_cases)
    
    logger.info(f"批量执行完成: 用户选择 {total_selected} 个，前置 {precondition_count} 个，跳过 {skipped_count} 个")
    logger.info(f"执行结果: 通过 {passed_count}, 失败 {failed_count}, 错误 {error_count}")
    
    return BatchExecuteResponse(
        total=total_selected,
        passed=passed_count,
        failed=failed_count,
        error=error_count,
        results=results
    )


# ========== 审批相关端点 ==========

@router.post("/cases/{case_id}/submit-review")
def submit_case_for_review(
    case_id: int,
    request: SubmitReviewRequest = SubmitReviewRequest(),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.TEST_CREATE)
):
    """提交API测试用例进行审批 (draft → pending_review)"""
    test_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"用例ID {case_id} 不存在")

    if test_case.status not in ("draft", "rejected"):
        raise HTTPException(status_code=400, detail=f"当前状态 '{test_case.status}' 不允许提交审批")

    test_case.status = "pending_review"
    test_case.reviewer_id = None
    test_case.review_comment = None
    test_case.reviewed_at = None
    test_case.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(test_case)

    logger.info(f"API测试用例 {case_id} 已提交审批 by user {current_user['user'].id}")

    return {"message": "提交审批成功", "case_id": case_id, "status": "pending_review"}


@router.post("/cases/{case_id}/review")
def review_test_case(
    case_id: int,
    request: ReviewActionRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.TEST_APPROVE)
):
    """审批API测试用例 (pending_review → approved/rejected)"""
    test_case = db.query(ApiTestCase).filter(ApiTestCase.id == case_id).first()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"用例ID {case_id} 不存在")

    if test_case.status != "pending_review":
        raise HTTPException(status_code=400, detail=f"当前状态 '{test_case.status}' 不允许审批，仅待审批状态的用例可审批")

    # 不允许自己审批自己创建的用例
    if test_case.created_by and test_case.created_by == str(current_user["user"].id):
        raise HTTPException(status_code=400, detail="不能审批自己创建的用例")

    new_status = "approved" if request.action == "approve" else "rejected"

    test_case.status = new_status
    test_case.reviewer_id = str(current_user["user"].id)
    test_case.review_comment = request.comment
    test_case.reviewed_at = datetime.utcnow()
    test_case.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(test_case)

    logger.info(f"API测试用例 {case_id} 已被 {request.action} by user {current_user['user'].id}")

    return {
        "message": f"审批{'通过' if request.action == 'approve' else '驳回'}成功",
        "case_id": case_id,
        "status": new_status,
        "reviewer_id": test_case.reviewer_id
    }


@router.get("/cases/review-statistics", response_model=ReviewStatisticsResponse)
def get_review_statistics(
    project_id: Optional[int] = Query(None, description="项目ID"),
    version_id: Optional[int] = Query(None, description="版本ID"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.TEST_READ)
):
    """获取API测试用例审批状态统计"""
    query = db.query(ApiTestCase)

    if project_id:
        query = query.filter(ApiTestCase.project_id == project_id)
    if version_id:
        query = query.filter(ApiTestCase.version_id == version_id)

    all_cases = query.all()

    stats = {
        "project_id": project_id or 0,
        "total": len(all_cases),
        "draft": sum(1 for c in all_cases if c.status == "draft"),
        "pending_review": sum(1 for c in all_cases if c.status == "pending_review"),
        "approved": sum(1 for c in all_cases if c.status == "approved"),
        "rejected": sum(1 for c in all_cases if c.status == "rejected"),
    }

    return stats


# ========== 导出端点 ==========

def _serialize_field(value: Any) -> str:
    """将字段值序列化为字符串（用于CSV导出）"""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


@router.get("/cases/export")
def export_test_cases(
    project_id: Optional[int] = Query(None, description="项目ID"),
    version_id: Optional[int] = Query(None, description="版本ID"),
    case_type: Optional[str] = Query(None, description="用例类型"),
    priority: Optional[str] = Query(None, description="优先级"),
    search: Optional[str] = Query(None, description="搜索关键词"),
    format: str = Query("csv", pattern="^(csv|xlsx)$", description="导出格式"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.TEST_READ)
):
    """导出API测试用例为CSV或Excel文件"""
    query = db.query(ApiTestCase)

    if project_id:
        query = query.filter(ApiTestCase.project_id == project_id)
    if version_id:
        query = query.filter(ApiTestCase.version_id == version_id)
    if case_type and case_type != "all":
        query = query.filter(ApiTestCase.case_type == case_type)
    if priority and priority != "all":
        query = query.filter(ApiTestCase.priority == priority)
    if search:
        query = query.filter(
            ApiTestCase.name.ilike(f"%{search}%") |
            ApiTestCase.path.ilike(f"%{search}%") |
            ApiTestCase.description.ilike(f"%{search}%")
        )

    cases = query.order_by(ApiTestCase.created_at.desc()).all()

    # 导出字段定义
    headers = [
        "ID", "用例名称", "请求方法", "请求路径", "用例类型", "优先级",
        "状态", "审批人", "审批意见", "基础URL", "请求头", "查询参数",
        "路径参数", "请求体", "预期状态码", "断言规则", "前置条件",
        "描述", "标签", "生成方式", "创建人", "创建时间", "更新时间"
    ]

    fields = [
        "id", "name", "method", "path", "case_type", "priority",
        "status", "reviewer_id", "review_comment", "base_url", "headers", "query_params",
        "path_params", "request_body", "expected_status", "assert_rules", "preconditions",
        "description", "tags", "generated_by", "created_by", "created_at", "updated_at"
    ]

    status_labels = {
        "draft": "草稿", "pending_review": "待审批",
        "approved": "已通过", "rejected": "已驳回"
    }

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)

        for case in cases:
            row = []
            for field in fields:
                if field == "status":
                    row.append(status_labels.get(getattr(case, field, ""), getattr(case, field, "")))
                elif field == "created_at" or field == "updated_at":
                    val = getattr(case, field, None)
                    row.append(val.isoformat() if val else "")
                else:
                    row.append(_serialize_field(getattr(case, field, None)))
            writer.writerow(row)

        csv_content = output.getvalue()
        output.close()

        # UTF-8 BOM for Excel compatibility
        csv_bytes = ('﻿' + csv_content).encode('utf-8')

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"API测试用例_{timestamp}.csv"

        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl 未安装，无法导出Excel格式")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "API测试用例"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, case in enumerate(cases, 2):
            for col_idx, field in enumerate(fields, 1):
                if field == "status":
                    value = status_labels.get(getattr(case, field, ""), getattr(case, field, ""))
                elif field == "created_at" or field == "updated_at":
                    val = getattr(case, field, None)
                    value = val.isoformat() if val else ""
                else:
                    value = _serialize_field(getattr(case, field, None))

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"API测试用例_{timestamp}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


# ========== 测试报告端点 ==========

@router.post("/executions/report", response_model=ReportResponse)
def generate_execution_report(
    request: ReportRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.TEST_READ)
):
    """生成API测试执行报告"""
    query = db.query(ApiTestExecution)

    if request.execution_ids:
        query = query.filter(ApiTestExecution.id.in_(request.execution_ids))
    elif request.project_id:
        query = query.filter(ApiTestExecution.project_id == request.project_id)
    else:
        raise HTTPException(status_code=400, detail="请提供 project_id 或 execution_ids")

    executions = query.order_by(ApiTestExecution.start_time.desc()).all()

    if not executions:
        raise HTTPException(status_code=404, detail="没有找到执行记录")

    # 统计
    total = len(executions)
    passed = sum(1 for e in executions if e.status == "passed")
    failed = sum(1 for e in executions if e.status == "failed")
    error = sum(1 for e in executions if e.status == "error")
    pass_rate = round(passed / total * 100, 2) if total > 0 else 0

    # 耗时统计
    durations = [e.duration for e in executions if e.duration is not None]
    duration_stats = {
        "avg_ms": round(sum(durations) / len(durations), 2) if durations else 0,
        "max_ms": max(durations) if durations else 0,
        "min_ms": min(durations) if durations else 0,
        "total_ms": sum(durations) if durations else 0,
    }

    # 按用例类型统计
    case_type_stats = {}
    for e in executions:
        case = db.query(ApiTestCase).filter(ApiTestCase.id == e.case_id).first()
        ct = case.case_type if case else "unknown"
        if ct not in case_type_stats:
            case_type_stats[ct] = {"passed": 0, "failed": 0, "total": 0}
        case_type_stats[ct]["total"] += 1
        if e.status == "passed":
            case_type_stats[ct]["passed"] += 1
        else:
            case_type_stats[ct]["failed"] += 1

    # 断言统计
    total_asserts = 0
    passed_asserts = 0
    for e in executions:
        if e.assert_results:
            for ar in (e.assert_results if isinstance(e.assert_results, list) else []):
                total_asserts += 1
                if ar.get("passed"):
                    passed_asserts += 1

    # 最慢用例
    slowest = sorted(
        [{"case_id": e.case_id, "duration": e.duration or 0,
          "name": (db.query(ApiTestCase).filter(ApiTestCase.id == e.case_id).first().name
                   if db.query(ApiTestCase).filter(ApiTestCase.id == e.case_id).first() else "Unknown")}
         for e in executions if e.duration],
        key=lambda x: x["duration"], reverse=True
    )[:5]

    # 结果详情
    results = []
    for e in executions:
        case = db.query(ApiTestCase).filter(ApiTestCase.id == e.case_id).first()
        results.append({
            "execution_id": e.id,
            "case_id": e.case_id,
            "case_name": case.name if case else "Unknown",
            "method": case.method if case else "",
            "path": case.path if case else "",
            "status": e.status,
            "duration": e.duration,
            "actual_status": e.actual_status,
            "error_message": e.error_message,
            "assert_results": e.assert_results if isinstance(e.assert_results, list) else [],
            "start_time": e.start_time.isoformat() if e.start_time else None,
        })

    return ReportResponse(
        project_id=request.project_id,
        version_id=request.version_id,
        report_time=datetime.utcnow(),
        total=total,
        passed=passed,
        failed=failed,
        error=error,
        pass_rate=pass_rate,
        duration_stats=duration_stats,
        case_type_stats=case_type_stats,
        results=results,
        assertion_summary={
            "total_asserts": total_asserts,
            "passed_asserts": passed_asserts,
            "failed_asserts": total_asserts - passed_asserts,
        },
        slowest_cases=slowest,
        most_failed_assertions=[],
    )


@router.get("/executions/report/export")
def export_report(
    project_id: int = Query(..., description="项目ID"),
    version_id: Optional[int] = Query(None, description="版本ID"),
    execution_ids: Optional[str] = Query(None, description="执行ID列表, 逗号分隔"),
    format: str = Query("html", pattern="^(html|pdf)$", description="导出格式"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.TEST_READ)
):
    """导出测试报告为HTML或PDF"""
    # 构建内部请求调用报告生成
    ids = [int(x.strip()) for x in execution_ids.split(",") if x.strip()] if execution_ids else None

    report_request = ReportRequest(
        project_id=project_id,
        version_id=version_id,
        execution_ids=ids,
    )
    report = generate_execution_report(report_request, db, current_user)

    if format == "html":
        html_content = _generate_html_report(report)
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            io.BytesIO(html_content.encode('utf-8')),
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=API测试报告_{timestamp}.html"}
        )
    elif format == "pdf":
        try:
            import weasyprint
            html_content = _generate_html_report(report)
            pdf_bytes = weasyprint.HTML(string=html_content).write_pdf()
        except ImportError:
            raise HTTPException(status_code=500, detail="weasyprint 未安装，无法导出PDF格式，请使用HTML格式")

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=API测试报告_{timestamp}.pdf"}
        )


def _generate_html_report(report: ReportResponse) -> str:
    """生成HTML格式的测试报告"""
    status_labels = {"passed": "通过", "failed": "失败", "error": "错误", "pending": "待执行", "running": "运行中"}
    status_colors = {"passed": "#52c41a", "failed": "#ff4d4f", "error": "#faad14", "pending": "#d9d9d9", "running": "#1890ff"}

    results_html = ""
    for r in report.results:
        color = status_colors.get(r["status"], "#d9d9d9")
        label = status_labels.get(r["status"], r["status"])
        results_html += f"""
        <tr>
            <td>{r['case_name']}</td>
            <td>{r.get('method', '')} {r.get('path', '')}</td>
            <td style="color:{color};font-weight:bold">{label}</td>
            <td>{r.get('duration', '-')}ms</td>
            <td>{r.get('actual_status', '-')}</td>
            <td style="max-width:300px;word-wrap:break-word">{r.get('error_message', '') or ''}</td>
        </tr>
        """

    case_type_stats_html = ""
    for ct, stats in report.case_type_stats.items():
        case_type_stats_html += f"""
        <tr>
            <td>{ct}</td>
            <td>{stats['total']}</td>
            <td style="color:#52c41a">{stats['passed']}</td>
            <td style="color:#ff4d4f">{stats['failed']}</td>
        </tr>
        """

    slowest_html = ""
    for i, s in enumerate(report.slowest_cases):
        slowest_html += f"<tr><td>{i+1}</td><td>{s['name']}</td><td>{s['duration']}ms</td></tr>"

    html = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>API测试报告</title>
    <style>
        body {{ font-family: 'Microsoft YaHei', 'PingFang SC', sans-serif; padding: 40px; color: #333; }}
        h1 {{ color: #1a1a2e; border-bottom: 3px solid #4472C4; padding-bottom: 10px; }}
        h2 {{ color: #4472C4; margin-top: 30px; }}
        .summary {{ display: flex; gap: 20px; margin: 20px 0; flex-wrap: wrap; }}
        .stat-card {{ flex: 1; min-width: 120px; padding: 20px; border-radius: 8px; text-align: center; color: #fff; }}
        .stat-card.total {{ background: #4472C4; }}
        .stat-card.passed {{ background: #52c41a; }}
        .stat-card.failed {{ background: #ff4d4f; }}
        .stat-card.rate {{ background: #722ed1; }}
        .stat-card .value {{ font-size: 32px; font-weight: bold; }}
        .stat-card .label {{ font-size: 14px; opacity: 0.9; margin-top: 5px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
        th, td {{ border: 1px solid #e8e8e8; padding: 10px 12px; text-align: left; font-size: 13px; }}
        th {{ background: #fafafa; font-weight: 600; }}
        tr:hover {{ background: #f5f5f5; }}
        .meta {{ color: #888; font-size: 13px; margin-bottom: 20px; }}
    </style>
</head>
<body>
    <h1>API 测试执行报告</h1>
    <p class="meta">生成时间: {report.report_time.strftime('%Y-%m-%d %H:%M:%S')} | 项目ID: {report.project_id}</p>

    <h2>概览</h2>
    <div class="summary">
        <div class="stat-card total">
            <div class="value">{report.total}</div>
            <div class="label">执行总数</div>
        </div>
        <div class="stat-card passed">
            <div class="value">{report.passed}</div>
            <div class="label">通过</div>
        </div>
        <div class="stat-card failed">
            <div class="value">{report.failed + report.error}</div>
            <div class="label">失败/错误</div>
        </div>
        <div class="stat-card rate">
            <div class="value">{report.pass_rate}%</div>
            <div class="label">通过率</div>
        </div>
    </div>

    <h2>耗时统计</h2>
    <table>
        <tr><th>指标</th><th>值</th></tr>
        <tr><td>平均耗时</td><td>{report.duration_stats.get('avg_ms', 0):.0f} ms</td></tr>
        <tr><td>最大耗时</td><td>{report.duration_stats.get('max_ms', 0)} ms</td></tr>
        <tr><td>最小耗时</td><td>{report.duration_stats.get('min_ms', 0)} ms</td></tr>
        <tr><td>总耗时</td><td>{report.duration_stats.get('total_ms', 0)} ms</td></tr>
    </table>

    <h2>断言统计</h2>
    <table>
        <tr><th>总断言数</th><th>通过</th><th>失败</th></tr>
        <tr>
            <td>{report.assertion_summary.get('total_asserts', 0)}</td>
            <td style="color:#52c41a">{report.assertion_summary.get('passed_asserts', 0)}</td>
            <td style="color:#ff4d4f">{report.assertion_summary.get('failed_asserts', 0)}</td>
        </tr>
    </table>

    <h2>按用例类型统计</h2>
    <table>
        <tr><th>类型</th><th>总数</th><th>通过</th><th>失败</th></tr>
        {case_type_stats_html}
    </table>

    <h2>最慢用例 TOP 5</h2>
    <table>
        <tr><th>#</th><th>用例名称</th><th>耗时</th></tr>
        {slowest_html}
    </table>

    <h2>执行明细</h2>
    <table>
        <tr><th>用例名称</th><th>接口</th><th>状态</th><th>耗时(ms)</th><th>状态码</th><th>错误信息</th></tr>
        {results_html}
    </table>
</body>
</html>
    """
    return html
def list_api_test_versions(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取API测试版本列表（包含项目版本和API测试专用版本）"""
    
    project_versions = db.query(Version).filter(Version.project_id == project_id).all()
    
    api_test_versions = db.query(ApiTestVersion).filter(ApiTestVersion.project_id == project_id).all()
    
    versions_list = []
    
    for pv in project_versions:
        existing_api_version = db.query(ApiTestVersion).filter(
            ApiTestVersion.project_id == project_id,
            ApiTestVersion.version_id == pv.id
        ).first()
        
        if not existing_api_version:
            api_version = ApiTestVersion(
                project_id=project_id,
                version_id=pv.id,
                name=pv.version_name or pv.version_number,
                version_number=pv.version_number,
                description=pv.description,
                is_api_test_only=False,
                created_at=pv.created_at
            )
            db.add(api_version)
            db.flush()
        
        api_version_record = existing_api_version or db.query(ApiTestVersion).filter(
            ApiTestVersion.project_id == project_id,
            ApiTestVersion.version_id == pv.id
        ).first()
        
        cases_count = db.query(ApiTestCase).filter(
            ApiTestCase.project_id == project_id,
            ApiTestCase.version_id == pv.id
        ).count()
        
        versions_list.append(ApiTestVersionResponse(
            id=api_version_record.id,
            project_id=project_id,
            version_id=pv.id,
            name=pv.version_name or pv.version_number,
            version_number=pv.version_number,
            description=pv.description,
            is_api_test_only=False,
            query_version_id=pv.id,
            test_cases_count=cases_count,
            created_by=api_version_record.created_by,
            created_at=api_version_record.created_at
        ))
    
    for av in api_test_versions:
        if av.is_api_test_only and av.version_id is None:
            cases_count = db.query(ApiTestCase).filter(
                ApiTestCase.project_id == project_id,
                ApiTestCase.version_id == av.id
            ).count()
            
            versions_list.append(ApiTestVersionResponse(
                id=av.id,
                project_id=project_id,
                version_id=av.version_id,
                name=av.name,
                version_number=av.version_number,
                description=av.description,
                is_api_test_only=True,
                query_version_id=av.id,
                test_cases_count=cases_count,
                created_by=av.created_by,
                created_at=av.created_at
            ))
    
    db.commit()
    
    versions_list.sort(key=lambda x: x.created_at, reverse=True)
    
    return ApiTestVersionListResponse(
        items=versions_list,
        total=len(versions_list)
    )


@router.get("/versions/{project_id}", response_model=ApiTestVersionListResponse)
def list_api_test_versions(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_READ)
):
    """获取API测试版本列表（包含项目版本和API测试专用版本）"""
    from app.core.models.project import Version

    project_versions = db.query(Version).filter(Version.project_id == project_id).all()

    api_test_versions = db.query(ApiTestVersion).filter(ApiTestVersion.project_id == project_id).all()

    versions_list = []

    for pv in project_versions:
        existing_api_version = db.query(ApiTestVersion).filter(
            ApiTestVersion.project_id == project_id,
            ApiTestVersion.version_id == pv.id
        ).first()

        if not existing_api_version:
            api_version = ApiTestVersion(
                project_id=project_id,
                version_id=pv.id,
                name=pv.version_name or pv.version_number,
                version_number=pv.version_number,
                description=pv.description,
                is_api_test_only=False,
                created_at=pv.created_at
            )
            db.add(api_version)
            db.flush()

        api_version_record = existing_api_version or db.query(ApiTestVersion).filter(
            ApiTestVersion.project_id == project_id,
            ApiTestVersion.version_id == pv.id
        ).first()

        cases_count = db.query(ApiTestCase).filter(
            ApiTestCase.project_id == project_id,
            ApiTestCase.version_id == pv.id
        ).count()

        versions_list.append(ApiTestVersionResponse(
            id=api_version_record.id,
            project_id=project_id,
            version_id=pv.id,
            name=pv.version_name or pv.version_number,
            version_number=pv.version_number,
            description=pv.description,
            is_api_test_only=False,
            query_version_id=pv.id,
            test_cases_count=cases_count,
            created_by=api_version_record.created_by,
            created_at=api_version_record.created_at
        ))

    for av in api_test_versions:
        if av.is_api_test_only and av.version_id is None:
            cases_count = db.query(ApiTestCase).filter(
                ApiTestCase.project_id == project_id,
                ApiTestCase.version_id == av.id
            ).count()

            versions_list.append(ApiTestVersionResponse(
                id=av.id,
                project_id=project_id,
                version_id=av.version_id,
                name=av.name,
                version_number=av.version_number,
                description=av.description,
                is_api_test_only=True,
                query_version_id=av.id,
                test_cases_count=cases_count,
                created_by=av.created_by,
                created_at=av.created_at
            ))

    db.commit()

    versions_list.sort(key=lambda x: x.created_at, reverse=True)

    return ApiTestVersionListResponse(
        items=versions_list,
        total=len(versions_list)
    )


@router.post("/versions", response_model=ApiTestVersionResponse, status_code=status.HTTP_201_CREATED)
def create_api_test_version(
    version_in: ApiTestVersionCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """创建API测试专用版本"""
    
    project = db.query(Project).filter(Project.id == version_in.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail=f"项目ID {version_in.project_id} 不存在")
    
    # 检查项目管理下是否存在同名版本（大小写不敏感）
    project_versions = db.query(Version).filter(Version.project_id == version_in.project_id).all()
    for pv in project_versions:
        if pv.version_name and pv.version_name.lower() == version_in.name.lower():
            raise HTTPException(status_code=400, detail=f"项目管理中已存在同名版本: {pv.version_name}，无需重复添加")
        if pv.version_number and version_in.version_number and pv.version_number.lower() == version_in.version_number.lower():
            raise HTTPException(status_code=400, detail=f"项目管理中已存在相同版本号: {pv.version_number}，无需重复添加")
        # 检查新输入的版本号是否与项目管理中的版本名称相同
        if version_in.version_number and pv.version_name and pv.version_name.lower() == version_in.version_number.lower():
            raise HTTPException(status_code=400, detail=f"版本号 {version_in.version_number} 与项目管理中的版本名称 {pv.version_name} 相同，请勿重复添加")
        # 检查新输入的版本名称是否与项目管理中的版本号相同
        if pv.version_number and pv.version_number.lower() == version_in.name.lower():
            raise HTTPException(status_code=400, detail=f"版本名称 {version_in.name} 与项目管理中的版本号 {pv.version_number} 相同，请勿重复添加")
    
    # 检查API测试版本表中是否存在同名版本（大小写不敏感）
    existing_api_versions = db.query(ApiTestVersion).filter(ApiTestVersion.project_id == version_in.project_id).all()
    for av in existing_api_versions:
        if av.name.lower() == version_in.name.lower():
            raise HTTPException(status_code=400, detail=f"已存在同名API测试版本: {av.name}")
        if av.version_number and version_in.version_number and av.version_number.lower() == version_in.version_number.lower():
            raise HTTPException(status_code=400, detail=f"已存在相同版本号的API测试版本: {av.version_number}")
        # 检查新输入的版本号是否与已存在的版本名称相同（大小写不敏感）
        if version_in.version_number and av.name.lower() == version_in.version_number.lower():
            raise HTTPException(status_code=400, detail=f"版本号 {version_in.version_number} 与已存在的版本名称 {av.name} 相同，请勿重复添加")
        # 检查新输入的版本名称是否与已存在的版本号相同（大小写不敏感）
        if av.version_number and av.version_number.lower() == version_in.name.lower():
            raise HTTPException(status_code=400, detail=f"版本名称 {version_in.name} 与已存在的版本号 {av.version_number} 相同，请勿重复添加")
    
    api_version = ApiTestVersion(
        project_id=version_in.project_id,
        version_id=None,
        name=version_in.name,
        version_number=version_in.version_number,
        description=version_in.description,
        is_api_test_only=True,
        created_by=str(current_user["user"].id),
        created_at=datetime.utcnow()
    )
    
    db.add(api_version)
    db.commit()
    db.refresh(api_version)
    
    logger.info(f"创建API测试版本: {api_version.name}")
    
    return ApiTestVersionResponse(
        id=api_version.id,
        project_id=api_version.project_id,
        version_id=api_version.version_id,
        name=api_version.name,
        version_number=api_version.version_number,
        description=api_version.description,
        is_api_test_only=api_version.is_api_test_only,
        query_version_id=api_version.id,
        test_cases_count=0,
        created_by=api_version.created_by,
        created_at=api_version.created_at
    )
    
    db.add(api_version)
    db.commit()
    db.refresh(api_version)
    
    logger.info(f"创建API测试版本: {api_version.name}")
    
    return ApiTestVersionResponse(
        id=api_version.id,
        project_id=api_version.project_id,
        version_id=api_version.version_id,
        name=api_version.name,
        version_number=api_version.version_number,
        description=api_version.description,
        is_api_test_only=api_version.is_api_test_only,
        query_version_id=api_version.id,
        test_cases_count=0,
        created_by=api_version.created_by,
        created_at=api_version.created_at
    )


@router.delete("/versions/{version_id}")
def delete_api_test_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE)
):
    """删除API测试专用版本（仅能删除is_api_test_only=True的版本）"""
    api_version = db.query(ApiTestVersion).filter(ApiTestVersion.id == version_id).first()
    if not api_version:
        raise HTTPException(status_code=404, detail=f"版本ID {version_id} 不存在")
    
    if not api_version.is_api_test_only:
        raise HTTPException(status_code=400, detail="不能删除项目同步的版本，请到项目管理中删除")
    
    db.delete(api_version)
    db.commit()
    
    logger.info(f"删除API测试专用版本: {version_id}")

# ===== 环境鉴权测试 =====

@router.post("/environments/{environment_id}/test-auth", response_model=TestAuthResponse)
async def test_environment_auth(
    environment_id: int,
    request: TestAuthRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(Permissions.VERSION_CREATE),
):
    """测试环境鉴权配置

    使用配置的鉴权信息尝试登录或验证凭证，返回结果。
    """
    env = db.query(ApiEnvironment).filter(
        ApiEnvironment.id == environment_id,
        ApiEnvironment.project_id == request.project_id
    ).first()

    if not env:
        raise HTTPException(status_code=404, detail="环境不存在")

    base_url = request.base_url or env.base_url or "http://localhost:8000"
    base_url = normalize_url_colon(base_url)

    result = await _execute_env_auth(env, base_url)

    token = result.get("auth_token")
    if token:
        # 脱敏：只显示前后4个字符
        if len(token) > 10:
            token_preview = token[:4] + "****" + token[-4:]
        else:
            token_preview = "****"

        return TestAuthResponse(
            success=True,
            message="鉴权成功，Token已获取",
            token_preview=token_preview,
            token_type="Bearer" if not env.auth_config.get("auth_type") in ["basic_auth", "api_key"] else env.auth_config.get("auth_type"),
        )
    elif result.get("auth_header"):
        return TestAuthResponse(
            success=True,
            message="Basic Auth凭证已生成",
        )
    elif result.get("api_keys"):
        return TestAuthResponse(
            success=True,
            message="API Key已配置",
        )
    elif result.get("cookies"):
        return TestAuthResponse(
            success=True,
            message="Cookie已配置",
        )
    else:
        error_msg = result.get("auth_error", "鉴权失败，未获取到有效凭证")
        return TestAuthResponse(
            success=False,
            message=error_msg,
        )


# ===== 文件Hash计算 =====

@router.post("/files/hash", response_model=FileHashResponse)
async def calculate_file_hash(
    file: UploadFile = File(...),
    current_user: dict = Depends(Permissions.VERSION_CREATE),
):
    """计算上传文件的哈希值(MD5/SHA1/SHA256)

    用于API测试中的文件内容校验、签名参数构造等场景。
    """
    import hashlib
    import mimetypes

    content = await file.read()

    result = {
        "file_name": file.filename,
        "file_size": len(content),
        "md5": hashlib.md5(content).hexdigest(),
        "sha1": hashlib.sha1(content).hexdigest(),
        "sha256": hashlib.sha256(content).hexdigest(),
    }

    # 推断MIME类型
    mime_type, _ = mimetypes.guess_type(file.filename or "")
    result["mime_type"] = mime_type or "application/octet-stream"

    logger.info(f"文件Hash计算完成: {file.filename}, size={len(content)}, md5={result['md5'][:8]}...")

    return FileHashResponse(**result)
    
    return {"message": "删除成功"}